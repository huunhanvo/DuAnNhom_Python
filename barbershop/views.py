from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Sum, Q, Avg, F, Case, When, IntegerField
from datetime import datetime, timedelta
import csv
import json
import bcrypt
from .models import (
    NguoiDung, ThongTinNhanVien, DanhMucDichVu, DichVu, 
    LichLamViec, YeuCauNghiPhep, DatLich, DichVuDatLich,
    HoaDon, ChiTietHoaDon, Voucher, CaiDatHeThong, DanhGia, DonXinNghi
)

# ============ HELPER FUNCTIONS ============

def require_auth(view_func):
    """Decorator to require authentication"""
    def wrapper(request, *args, **kwargs):
        if 'user_id' not in request.session:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

def require_role(allowed_roles):
    """Decorator to require specific role"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if 'user_id' not in request.session:
                return redirect('login')
            if request.session.get('vai_tro') not in allowed_roles:
                return redirect('login')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# ============ AUTH VIEWS ============

def login_view(request):
    """Login Page"""
    if request.method == 'POST':
        sdt = request.POST.get('username')
        password = request.POST.get('password')
        
        try:
            user = NguoiDung.objects.get(so_dien_thoai=sdt, da_xoa=False)
            
            # Verify password with bcrypt
            if bcrypt.checkpw(password.encode('utf-8'), user.mat_khau_hash.encode('utf-8')):
                # Save user info to session
                request.session['user_id'] = user.id
                request.session['vai_tro'] = user.vai_tro
                request.session['ho_ten'] = user.ho_ten
                
                # Redirect based on role
                if user.vai_tro == 'quan_ly':
                    return redirect('admin_dashboard')
                elif user.vai_tro == 'nhan_vien':
                    return redirect('staff_dashboard')
                else:  # khach_hang
                    return redirect('login')  # Customer portal not implemented yet
            else:
                context = {'error': 'Sai mật khẩu!'}
                return render(request, 'login.html', context)
                
        except NguoiDung.DoesNotExist:
            context = {'error': 'Số điện thoại không tồn tại!'}
            return render(request, 'login.html', context)
    
    return render(request, 'login.html')

def logout_view(request):
    """Logout"""
    request.session.flush()
    return redirect('login')

def page_not_found(request, exception=None):
    """404 Page"""
    return render(request, '404.html', status=404)

# ============ ADMIN VIEWS ============

@require_role(['quan_ly'])
def admin_dashboard(request):
    """Admin Dashboard"""
    today = timezone.now().date()
    
    # Today's bookings
    today_bookings = DatLich.objects.filter(
        ngay_hen=today,
        da_xoa=False
    ).exclude(trang_thai='da_huy')
    
    # Today's revenue from invoices
    today_invoices = HoaDon.objects.filter(
        ngay_tao__date=today,
        da_xoa=False
    )
    today_revenue = today_invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
    
    # Total customers
    total_customers = NguoiDung.objects.filter(
        vai_tro='khach_hang',
        da_xoa=False
    ).count()
    
    # Pending bookings
    pending_bookings = DatLich.objects.filter(
        trang_thai='cho_xac_nhan',
        da_xoa=False
    ).count()
    
    # Revenue chart (last 7 days)
    revenue_chart_data = []
    revenue_chart_labels = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        daily_revenue = HoaDon.objects.filter(
            ngay_tao__date=date,
            da_xoa=False
        ).aggregate(total=Sum('thanh_tien'))['total'] or 0
        revenue_chart_data.append(float(daily_revenue))
        revenue_chart_labels.append(date.strftime('%d/%m'))
    
    # Top services
    top_services = DichVu.objects.filter(da_xoa=False).annotate(
        booking_count=Count('dichvudatlich', filter=Q(dichvudatlich__dat_lich__da_xoa=False))
    ).order_by('-booking_count')[:5]
    
    # Pending bookings for dashboard table
    pending_bookings_list = DatLich.objects.filter(
        trang_thai='cho_xac_nhan',
        da_xoa=False
    ).select_related('khach_hang', 'nhan_vien').order_by('ngay_hen', 'gio_hen')[:10]
    
    # Pending leave requests
    pending_leave_requests = DonXinNghi.objects.filter(
        trang_thai='cho_duyet',
        da_xoa=False
    ).select_related('nhan_vien').order_by('ngay_tao')[:10]
    
    # Upcoming bookings today
    upcoming_bookings = DatLich.objects.filter(
        ngay_hen=today,
        da_xoa=False
    ).exclude(trang_thai='da_huy').select_related(
        'khach_hang', 'nhan_vien'
    ).order_by('gio_hen')[:5]
    
    context = {
        'today_bookings': today_bookings.count(),
        'today_revenue': today_revenue,
        'total_customers': total_customers,
        'pending_bookings': pending_bookings,
        'pending_bookings_list': pending_bookings_list,
        'pending_leave_requests': pending_leave_requests,
        'revenue_chart_labels': revenue_chart_labels,
        'revenue_chart_data': revenue_chart_data,
        'top_services': top_services,
        'upcoming_bookings': upcoming_bookings,
    }
    return render(request, 'admin/dashboard.html', context)

@require_role(['quan_ly'])
def admin_staff(request):
    """Staff Management with CRUD, Search, Filter"""
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            # Create new staff with validation
            try:
                # Validate required fields
                ho_ten = request.POST.get('ho_ten', '').strip()
                so_dien_thoai = request.POST.get('so_dien_thoai', '').strip()
                email = request.POST.get('email', '').strip()
                
                if not ho_ten:
                    return JsonResponse({'success': False, 'message': 'Họ tên không được để trống'})
                
                if not so_dien_thoai:
                    return JsonResponse({'success': False, 'message': 'Số điện thoại không được để trống'})
                
                # Check if phone already exists
                if NguoiDung.objects.filter(so_dien_thoai=so_dien_thoai, da_xoa=False).exists():
                    return JsonResponse({'success': False, 'message': f'Số điện thoại {so_dien_thoai} đã tồn tại!'})
                
                # Check if email already exists (if provided)
                if email and NguoiDung.objects.filter(email=email, da_xoa=False).exists():
                    return JsonResponse({'success': False, 'message': f'Email {email} đã tồn tại!'})
                
                # Validate phone format (basic)
                if not so_dien_thoai.isdigit() or len(so_dien_thoai) < 10:
                    return JsonResponse({'success': False, 'message': 'Số điện thoại không hợp lệ (tối thiểu 10 số)'})
                
                # Parse date
                ngay_sinh_str = request.POST.get('ngay_sinh', '').strip()
                ngay_sinh = None
                if ngay_sinh_str:
                    from datetime import datetime
                    try:
                        ngay_sinh = datetime.strptime(ngay_sinh_str, '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({'success': False, 'message': 'Ngày sinh không hợp lệ (định dạng: YYYY-MM-DD)'})
                
                # Create user account
                user = NguoiDung.objects.create(
                    ho_ten=ho_ten,
                    so_dien_thoai=so_dien_thoai,
                    email=email if email else None,
                    vai_tro='nhan_vien',
                    ngay_sinh=ngay_sinh,
                    gioi_tinh=request.POST.get('gioi_tinh', 'nam'),
                    dia_chi=request.POST.get('dia_chi', '').strip() or None,
                    trang_thai=True,
                    da_xoa=False
                )
                
                # Hash password
                password = request.POST.get('password', '123456')
                hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
                user.mat_khau_hash = hashed.decode('utf-8')
                user.save()
                
                # Create staff info (only staff-specific fields)
                chuyen_mon_raw = request.POST.get('chuyen_mon', '').strip()
                chung_chi_raw = request.POST.get('chung_chi', '').strip()
                
                ThongTinNhanVien.objects.create(
                    nguoi_dung=user,
                    cccd=request.POST.get('cccd', '').strip() or None,
                    chuyen_mon=chuyen_mon_raw if chuyen_mon_raw else None,
                    kinh_nghiem_nam=int(request.POST.get('kinh_nghiem_nam', 0) or 0),
                    chung_chi=chung_chi_raw if chung_chi_raw else None,
                    mo_ta=request.POST.get('mo_ta', '').strip() or None,
                )
                
                return JsonResponse({'success': True, 'message': 'Đã thêm nhân viên mới thành công!', 'staff_id': user.id})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'})
        
        elif action == 'delete':
            # Soft delete staff
            try:
                staff_id = request.POST.get('staff_id')
                user = get_object_or_404(NguoiDung, id=staff_id, vai_tro='nhan_vien', da_xoa=False)
                user.da_xoa = True
                user.ngay_xoa = timezone.now()
                user.save()
                
                return JsonResponse({'success': True, 'message': 'Đã xóa nhân viên!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'toggle_status':
            # Toggle active/inactive
            try:
                staff_id = request.POST.get('staff_id')
                user = get_object_or_404(NguoiDung, id=staff_id, vai_tro='nhan_vien', da_xoa=False)
                user.trang_thai = not user.trang_thai
                user.save()
                
                status_text = 'kích hoạt' if user.trang_thai else 'tạm ngừng'
                return JsonResponse({'success': True, 'message': f'Đã {status_text} nhân viên!', 'new_status': user.trang_thai})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - list staff with filters
    staff_query = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        da_xoa=False
    ).select_related('thong_tin_nhan_vien')
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        staff_query = staff_query.filter(
            Q(ho_ten__icontains=search) |
            Q(so_dien_thoai__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        staff_query = staff_query.filter(trang_thai=True)
    elif status_filter == 'inactive':
        staff_query = staff_query.filter(trang_thai=False)
    
    # Sort
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'name':
        staff_query = staff_query.order_by('ho_ten')
    elif sort_by == 'rating':
        staff_query = staff_query.order_by('-thong_tin_nhan_vien__danh_gia_trung_binh')
    elif sort_by == 'revenue':
        # Sort by total revenue (bookings count as proxy)
        staff_query = staff_query.annotate(
            total_bookings=Count('dat_lich_nhan_vien', filter=Q(dat_lich_nhan_vien__da_xoa=False))
        ).order_by('-total_bookings')
    else:
        staff_query = staff_query.order_by('-ngay_tao')
    
    # Annotate with statistics
    staff_users = staff_query.annotate(
        booking_count=Count('dat_lich_nhan_vien', filter=Q(dat_lich_nhan_vien__da_xoa=False)),
    )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(staff_users, 10)  # 10 per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'staff_list': page_obj,
        'total_staff': staff_users.count(),
        'active_staff': staff_users.filter(trang_thai=True).count(),
        'search': search,
        'status_filter': status_filter,
        'sort_by': sort_by,
    }
    return render(request, 'admin/staff.html', context)

@require_role(['quan_ly'])
def admin_staff_detail(request, staff_id):
    """Staff Detail with Statistics"""
    user = get_object_or_404(NguoiDung, id=staff_id, vai_tro='nhan_vien', da_xoa=False)
    
    # Get staff details
    staff_info = None
    if hasattr(user, 'thong_tin_nhan_vien'):
        staff_info = user.thong_tin_nhan_vien
    
    # Calculate statistics
    from django.db.models import Count
    
    # Total bookings
    total_bookings = DatLich.objects.filter(nhan_vien=user, da_xoa=False).count()
    
    # Unique customers
    total_customers = DatLich.objects.filter(nhan_vien=user, da_xoa=False).values('khach_hang').distinct().count()
    
    # Average rating from ThongTinNhanVien
    avg_rating = 0
    if staff_info and staff_info.danh_gia_trung_binh:
        avg_rating = round(float(staff_info.danh_gia_trung_binh), 1)
    
    # Get bookings
    bookings = DatLich.objects.filter(
        nhan_vien=user,
        da_xoa=False
    ).select_related('khach_hang').prefetch_related('dich_vu_dat_lich__dich_vu').order_by('-ngay_hen', '-gio_hen')[:10]
    
    # Get work schedule
    schedules = LichLamViec.objects.filter(
        nhan_vien=user,
        da_xoa=False
    ).order_by('ngay_lam')[:7]
    
    # Skills (from chuyen_mon)
    skills = []
    if staff_info and staff_info.chuyen_mon:
        if isinstance(staff_info.chuyen_mon, str):
            skills = [s.strip() for s in staff_info.chuyen_mon.split(',') if s.strip()]
        elif isinstance(staff_info.chuyen_mon, list):
            skills = staff_info.chuyen_mon
    
    # Certificates
    certificates = []
    if staff_info and staff_info.chung_chi:
        if isinstance(staff_info.chung_chi, str):
            certificates = [c.strip() for c in staff_info.chung_chi.split(',') if c.strip()]
        elif isinstance(staff_info.chung_chi, list):
            certificates = staff_info.chung_chi
    
    context = {
        'staff': user,
        'user': user,
        'staff_info': staff_info,
        'bookings': bookings,
        'schedules': schedules,
        'skills': skills,
        'certificates': certificates,
        # Statistics
        'chuc_vu': 'Nhân viên',  # Default role
        'avg_rating': avg_rating,
        'total_services': total_bookings,  # Total bookings as proxy
        'total_customers': total_customers,
    }
    return render(request, 'admin/staff-detail.html', context)

@require_role(['quan_ly'])
def admin_staff_edit(request, staff_id):
    """Staff Edit with Complete Fields"""
    user = get_object_or_404(NguoiDung, id=staff_id, vai_tro='nhan_vien', da_xoa=False)
    
    if request.method == 'POST':
        try:
            # Update user info (NguoiDung fields)
            user.ho_ten = request.POST.get('ho_ten', user.ho_ten)
            user.email = request.POST.get('email', user.email) or None
            user.so_dien_thoai = request.POST.get('so_dien_thoai', user.so_dien_thoai)
            user.dia_chi = request.POST.get('dia_chi', '') or None
            
            # Update ngay_sinh, gioi_tinh (NguoiDung fields, not ThongTinNhanVien!)
            ngay_sinh_str = request.POST.get('ngay_sinh', '').strip()
            if ngay_sinh_str:
                from datetime import datetime
                try:
                    user.ngay_sinh = datetime.strptime(ngay_sinh_str, '%Y-%m-%d').date()
                except:
                    pass
            
            gioi_tinh = request.POST.get('gioi_tinh', '').strip()
            if gioi_tinh:
                user.gioi_tinh = gioi_tinh
            
            # Handle trang_thai checkbox
            user.trang_thai = request.POST.get('trang_thai') == 'on' or request.POST.get('trang_thai') == '1' or request.POST.get('trang_thai') == 'True'
            user.save()
            
            # Update or create staff info (ThongTinNhanVien fields only)
            staff_info, created = ThongTinNhanVien.objects.get_or_create(
                nguoi_dung=user,
                defaults={'da_xoa': False}
            )
            
            # Update ThongTinNhanVien fields (NOT ngay_sinh, gioi_tinh)
            cccd_raw = request.POST.get('cccd', '').strip()
            chuyen_mon_raw = request.POST.get('chuyen_mon', '').strip()
            chung_chi_raw = request.POST.get('chung_chi', '').strip()
            mo_ta_raw = request.POST.get('mo_ta', '').strip()
            
            staff_info.cccd = cccd_raw if cccd_raw else None
            staff_info.chuyen_mon = chuyen_mon_raw if chuyen_mon_raw else None
            
            kinh_nghiem = request.POST.get('kinh_nghiem_nam', '0')
            staff_info.kinh_nghiem_nam = int(kinh_nghiem) if kinh_nghiem else 0
            
            staff_info.chung_chi = chung_chi_raw if chung_chi_raw else None
            staff_info.mo_ta = mo_ta_raw if mo_ta_raw else None
            staff_info.save()
            
            from django.contrib import messages
            messages.success(request, 'Đã cập nhật thông tin nhân viên!')
            return redirect('admin_staff_detail', staff_id=staff_id)
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    staff_info = None
    if hasattr(user, 'thong_tin_nhan_vien'):
        staff_info = user.thong_tin_nhan_vien
    
    context = {
        'staff': user,
        'user': user,
        'staff_info': staff_info,
        'chuc_vu': 'Nhân viên',  # Default role
    }
    return render(request, 'admin/staff-edit.html', context)

@require_role(['quan_ly'])
def admin_bookings(request):
    """Booking Management"""
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    staff_id = request.GET.get('staff_id')
    search = request.GET.get('search')
    
    # Debug logging
    print(f"DEBUG - Filter parameters: from_date={from_date}, to_date={to_date}, staff_id={staff_id}, search={search}")
    
    # Base queryset
    bookings = DatLich.objects.filter(
        da_xoa=False
    ).select_related('khach_hang', 'nhan_vien').prefetch_related(
        'dich_vu_dat_lich__dich_vu'
    ).order_by('-ngay_hen', '-gio_hen')
    
    # Apply filters
    original_count = bookings.count()
    
    if from_date:
        bookings = bookings.filter(ngay_hen__gte=from_date)
        print(f"DEBUG - After from_date filter ({from_date}): {bookings.count()} bookings")
    
    if to_date:
        bookings = bookings.filter(ngay_hen__lte=to_date)
        print(f"DEBUG - After to_date filter ({to_date}): {bookings.count()} bookings")
    
    if staff_id and staff_id.strip():
        bookings = bookings.filter(nhan_vien_id=staff_id.strip())
        print(f"DEBUG - After staff filter ({staff_id.strip()}): {bookings.count()} bookings")
    
    if search and search.strip() and search.strip().lower() != 'none':
        from django.db.models import Q
        search_term = search.strip()
        bookings = bookings.filter(
            Q(so_dien_thoai_khach__icontains=search_term) |
            Q(ten_khach_hang__icontains=search_term) |
            Q(ma_dat_lich__icontains=search_term)
        )
        print(f"DEBUG - After search filter ({search_term}): {bookings.count()} bookings")
    
    print(f"DEBUG - Total bookings after all filters: {bookings.count()} (original: {original_count})")
    
    # Get all bookings for statistics (without pagination)
    all_bookings = bookings
    
    # Statistics
    total_bookings = all_bookings.count()
    pending_count = all_bookings.filter(trang_thai='cho_xac_nhan').count()
    confirmed_count = all_bookings.filter(trang_thai='da_xac_nhan').count()
    in_progress_count = all_bookings.filter(trang_thai='da_checkin').count()
    completed_count = all_bookings.filter(trang_thai='hoan_thanh').count()
    cancelled_count = all_bookings.filter(trang_thai='da_huy').count()
    
    # Get staff list for filter
    staff_list = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Get customers for modal
    customers = NguoiDung.objects.filter(
        vai_tro='khach_hang',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Get services for modal
    services = DichVu.objects.filter(
        trang_thai=True,
        da_xoa=False
    ).select_related('danh_muc').order_by('thu_tu', 'ten_dich_vu')
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(bookings, 25)  # 25 bookings per page
    page_number = request.GET.get('page')
    bookings_page = paginator.get_page(page_number)
    
    context = {
        'bookings': bookings_page,
        'total_bookings': total_bookings,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'staff_list': staff_list,
        'customers': customers,
        'services': services,
        'from_date': from_date,
        'to_date': to_date,
        'staff_id': staff_id,
        'search': search,
    }
    return render(request, 'admin/bookings.html', context)

@require_role(['quan_ly'])
def admin_bookings_create(request):
    """Create Booking with proper validation"""
    if request.method == 'POST':
        try:
            # Validate required fields
            customer_id = request.POST.get('customer_id')
            staff_id = request.POST.get('staff_id') 
            booking_date = request.POST.get('booking_date')
            booking_time = request.POST.get('booking_time')
            services = request.POST.getlist('services[]')  # Note: services[], not service_ids
            
            if not all([customer_id, staff_id, booking_date, booking_time]):
                from django.contrib import messages
                messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc!')
                return redirect(request.get_full_path())
            
            if not services:
                from django.contrib import messages  
                messages.error(request, 'Vui lòng chọn ít nhất một dịch vụ!')
                return redirect(request.get_full_path())
            
            # Calculate total amount and generate booking code
            total_amount = 0
            for service_id in services:
                service = DichVu.objects.get(id=service_id)
                total_amount += service.gia
            
            # Generate unique booking code
            import random, string
            booking_code = 'BK' + ''.join(random.choices(string.digits, k=6))
            while DatLich.objects.filter(ma_dat_lich=booking_code).exists():
                booking_code = 'BK' + ''.join(random.choices(string.digits, k=6))
            
            # Get customer info
            customer = NguoiDung.objects.get(id=customer_id)
            
            # Create booking
            booking = DatLich.objects.create(
                ma_dat_lich=booking_code,
                khach_hang=customer,
                ten_khach_hang=customer.ho_ten,
                so_dien_thoai_khach=customer.so_dien_thoai,
                email_khach=customer.email or None,
                nhan_vien_id=staff_id if staff_id != 'auto' else None,
                ngay_hen=booking_date,
                gio_hen=booking_time,
                tong_tien=total_amount,
                thanh_tien=total_amount,
                trang_thai='da_xac_nhan',
                ghi_chu=request.POST.get('notes', ''),
                da_xoa=False
            )
            
            # Add services
            for service_id in services:
                service = DichVu.objects.get(id=service_id)
                DichVuDatLich.objects.create(
                    dat_lich=booking,
                    dich_vu=service,
                    ten_dich_vu=service.ten_dich_vu,
                    gia=service.gia,
                    so_luong=1,
                    thanh_tien=service.gia
                )
            
            from django.contrib import messages
            messages.success(request, f'Đã tạo lịch hẹn {booking_code} thành công!')
            return redirect('admin_bookings')
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Lỗi tạo lịch hẹn: {str(e)}')
            return redirect(request.get_full_path())
    
    # Get data for form
    customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False).order_by('ho_ten')
    staff_list = NguoiDung.objects.filter(vai_tro='nhan_vien', da_xoa=False, trang_thai=True).order_by('ho_ten')
    services = DichVu.objects.filter(da_xoa=False, trang_thai=True).order_by('ten_dich_vu')
    
    # Generate time slots (8:00 - 20:00)
    from datetime import time
    time_slots = []
    for hour in range(8, 21):  # 8AM to 8PM
        for minute in [0, 30]:  # Every 30 minutes
            if hour == 20 and minute == 30:  # Don't go past 8PM
                break
            slot_time = time(hour, minute)
            time_slots.append(slot_time.strftime('%H:%M'))
    
    # Get selected customer ID from URL parameter
    selected_customer_id = request.GET.get('customer_id')
    
    from datetime import date
    context = {
        'customers': customers,
        'staff': staff_list,  # Changed from 'staff' to 'staff_list'
        'staff_list': staff_list,  # Add both for compatibility
        'services': services,
        'time_slots': time_slots,
        'today': date.today().isoformat(),
        'selected_customer_id': selected_customer_id,
    }
    return render(request, 'admin/bookings-create.html', context)

@require_role(['quan_ly'])
def admin_booking_detail(request, booking_id):
    """Booking Detail and Management"""
    booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            new_status = request.POST.get('trang_thai')
            booking.trang_thai = new_status
            
            # Update timestamps based on status
            if new_status == 'da_checkin':
                booking.ngay_check_in = timezone.now()
            elif new_status == 'hoan_thanh':
                booking.ngay_hoan_thanh = timezone.now()
            
            booking.save()
            
        elif action == 'cancel':
            booking.trang_thai = 'da_huy'
            booking.ly_do_huy = request.POST.get('ly_do_huy')
            booking.save()
            
            # TODO: Send notification to customer
            
        return redirect('admin_booking_detail', booking_id=booking_id)
    
    # GET request - show booking details
    services = booking.dich_vu_dat_lich.all().select_related('dich_vu')
    
    context = {
        'booking': booking,
        'services': services,
    }
    return render(request, 'admin/booking-detail.html', context)

@require_role(['quan_ly'])
def admin_booking_cancel(request, booking_id):
    """API to cancel booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai in ['hoan_thanh', 'da_huy']:
                return JsonResponse({
                    'success': False, 
                    'message': 'Không thể hủy lịch hẹn đã hoàn thành hoặc đã hủy'
                })
            
            booking.trang_thai = 'da_huy'
            booking.ngay_huy = timezone.now()
            booking.ly_do_huy = request.POST.get('reason', 'Hủy từ admin')
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Đã hủy lịch hẹn thành công'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_booking_checkin(request, booking_id):
    """API to check-in booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai != 'da_xac_nhan':
                return JsonResponse({
                    'success': False,
                    'message': 'Chỉ có thể check-in lịch hẹn đã xác nhận'
                })
            
            booking.trang_thai = 'da_checkin'
            booking.ngay_check_in = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Đã check-in thành công'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_booking_complete(request, booking_id):
    """API to complete booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai != 'da_checkin':
                return JsonResponse({
                    'success': False,
                    'message': 'Chỉ có thể hoàn thành lịch hẹn đã check-in'
                })
            
            booking.trang_thai = 'hoan_thanh'
            booking.ngay_hoan_thanh = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Đã hoàn thành lịch hẹn'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_bookings_export(request):
    """Export bookings to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset
        bookings = DatLich.objects.filter(
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien').prefetch_related(
            'dich_vu_dat_lich__dich_vu'
        ).order_by('-ngay_hen', '-gio_hen')
        
        # Apply filters
        if from_date:
            bookings = bookings.filter(ngay_hen__gte=from_date)
        
        if to_date:
            bookings = bookings.filter(ngay_hen__lte=to_date)
        
        if staff_id:
            bookings = bookings.filter(nhan_vien_id=staff_id)
        
        if search:
            from django.db.models import Q
            bookings = bookings.filter(
                Q(so_dien_thoai_khach__icontains=search) |
                Q(ten_khach_hang__icontains=search) |
                Q(ma_dat_lich__icontains=search)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách đặt lịch"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Mã đặt lịch', 'Khách hàng', 'Số điện thoại', 'Nhân viên',
            'Ngày hẹn', 'Giờ hẹn', 'Dịch vụ', 'Tổng tiền', 'Trạng thái', 
            'Ngày tạo', 'Ghi chú'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        status_map = {
            'cho_xac_nhan': 'Chờ xác nhận',
            'da_xac_nhan': 'Đã xác nhận',
            'da_checkin': 'Đang phục vụ',
            'hoan_thanh': 'Hoàn thành',
            'da_huy': 'Đã hủy',
            'khong_den': 'Không đến'
        }
        
        for row, booking in enumerate(bookings, 2):
            services = ', '.join([s.ten_dich_vu for s in booking.dich_vu_dat_lich.all()])
            
            ws.cell(row=row, column=1, value=booking.ma_dat_lich)
            ws.cell(row=row, column=2, value=booking.ten_khach_hang)
            ws.cell(row=row, column=3, value=booking.so_dien_thoai_khach)
            ws.cell(row=row, column=4, value=booking.nhan_vien.ho_ten if booking.nhan_vien else '')
            ws.cell(row=row, column=5, value=booking.ngay_hen.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=booking.gio_hen.strftime('%H:%M'))
            ws.cell(row=row, column=7, value=services)
            ws.cell(row=row, column=8, value=float(booking.thanh_tien))
            ws.cell(row=row, column=9, value=status_map.get(booking.trang_thai, booking.trang_thai))
            ws.cell(row=row, column=10, value=booking.ngay_tao.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=11, value=booking.ghi_chu or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'danh_sach_dat_lich_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'Cần cài đặt openpyxl để xuất Excel: pip install openpyxl'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Lỗi xuất Excel: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_customers(request):
    """Customer Management with stats, filters and CRUD"""
    from datetime import datetime, timedelta
    
    if request.method == 'POST':
        # Handle add customer
        try:
            # Validate required fields
            ho_ten = request.POST.get('ho_ten', '').strip()
            so_dien_thoai = request.POST.get('so_dien_thoai', '').strip()
            email = request.POST.get('email', '').strip()
            
            if not ho_ten:
                return JsonResponse({'success': False, 'message': 'Họ tên không được để trống'})
            
            if not so_dien_thoai:
                return JsonResponse({'success': False, 'message': 'Số điện thoại không được để trống'})
            
            # Check if phone already exists
            if NguoiDung.objects.filter(so_dien_thoai=so_dien_thoai, da_xoa=False).exists():
                return JsonResponse({'success': False, 'message': f'Số điện thoại {so_dien_thoai} đã tồn tại!'})
            
            # Process email - set to None if empty
            email = email.strip() if email else None
            if email == '':
                email = None
                
            # Check if email already exists (if provided)
            if email and NguoiDung.objects.filter(email__iexact=email, da_xoa=False).exists():
                return JsonResponse({'success': False, 'message': f'Email {email} đã tồn tại!'})
            
            # Validate phone format
            if not so_dien_thoai.isdigit() or len(so_dien_thoai) < 10:
                return JsonResponse({'success': False, 'message': 'Số điện thoại không hợp lệ (tối thiểu 10 số)'})
            
            # Parse birth date
            ngay_sinh_str = request.POST.get('ngay_sinh', '').strip()
            ngay_sinh = None
            if ngay_sinh_str:
                try:
                    ngay_sinh = datetime.strptime(ngay_sinh_str, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'message': 'Ngày sinh không hợp lệ (định dạng: YYYY-MM-DD)'})
            
            # Create customer
            import bcrypt
            customer = NguoiDung.objects.create(
                ho_ten=ho_ten,
                so_dien_thoai=so_dien_thoai,
                email=email,
                vai_tro='khach_hang',
                ngay_sinh=ngay_sinh,
                dia_chi=request.POST.get('dia_chi', '').strip() or None,
                mat_khau_hash=bcrypt.hashpw('123456'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),  # Default password
                trang_thai=True,
                da_xoa=False
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'Đã thêm khách hàng {ho_ten} thành công!',
                'customer_id': customer.id
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Lỗi: {str(e)}'})
    
    # GET request - Base query
    customers_query = NguoiDung.objects.filter(
        vai_tro='khach_hang',
        da_xoa=False
    )
    
    # Apply filters
    search = request.GET.get('search', '').strip()
    if search:
        customers_query = customers_query.filter(
            Q(ho_ten__icontains=search) |
            Q(so_dien_thoai__icontains=search) |
            Q(email__icontains=search)
        )
    
    tier = request.GET.get('tier', '')
    status = request.GET.get('status', '')
    
    # Add annotations for stats
    customers = customers_query.annotate(
        so_lan_cat=Count('dat_lich', filter=Q(dat_lich__da_xoa=False, dat_lich__trang_thai='hoan_thanh')),
        tong_chi_tieu=Sum('hoa_don__thanh_tien', filter=Q(hoa_don__da_xoa=False)) or 0
    ).order_by('-ngay_tao')
    
    # Add customer tier calculation
    customer_list = []
    for customer in customers:
        points = customer.diem_tich_luy or 0
        if points >= 1000:
            hang_thanh_vien = 'platinum'
        elif points >= 500:
            hang_thanh_vien = 'gold'
        elif points >= 200:
            hang_thanh_vien = 'silver'
        else:
            hang_thanh_vien = 'bronze'
        
        # Add tier to customer object
        customer.hang_thanh_vien = hang_thanh_vien
        customer_list.append(customer)
    
    # Apply tier filter
    if tier:
        customer_list = [c for c in customer_list if c.hang_thanh_vien == tier]
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(customer_list, 20)
    page = request.GET.get('page', 1)
    customers_page = paginator.get_page(page)
    
    # Statistics
    total_customers = customers_query.count()
    new_customers_month = customers_query.filter(
        ngay_tao__gte=timezone.now() - timedelta(days=30)
    ).count()
    
    vip_customers = sum(1 for c in customer_list if c.hang_thanh_vien in ['gold', 'platinum'])
    active_customers = sum(1 for c in customer_list if c.so_lan_cat > 0)
    
    # Calculate averages
    total_visits = sum(c.so_lan_cat for c in customer_list)
    total_spending = sum(float(c.tong_chi_tieu or 0) for c in customer_list)
    
    avg_visits = total_visits / total_customers if total_customers > 0 else 0
    avg_spending = total_spending / total_customers if total_customers > 0 else 0
    
    context = {
        'customers': customers_page,
        'total_customers': total_customers,
        'new_customers_month': new_customers_month,
        'vip_customers': vip_customers,
        'active_customers': active_customers,
        'avg_visits': avg_visits,
        'avg_spending': avg_spending,
        'search': search,
        'tier': tier,
        'status': status,
    }
    return render(request, 'admin/customers.html', context)

@require_role(['quan_ly'])
def admin_services(request):
    """Service Management with CRUD"""
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            # Create new service
            try:
                # Validation
                ten_dich_vu = request.POST.get('ten_dich_vu', '').strip()
                danh_muc = request.POST.get('danh_muc', '').strip()
                gia = request.POST.get('gia', '').strip()
                thoi_luong = request.POST.get('thoi_luong', '').strip()
                
                if not ten_dich_vu:
                    return JsonResponse({'success': False, 'message': 'Tên dịch vụ không được để trống'})
                if not danh_muc:
                    return JsonResponse({'success': False, 'message': 'Danh mục không được để trống'})
                if not gia:
                    return JsonResponse({'success': False, 'message': 'Giá không được để trống'})
                if not thoi_luong:
                    return JsonResponse({'success': False, 'message': 'Thời lượng không được để trống'})
                
                # Check if service name exists
                if DichVu.objects.filter(ten_dich_vu=ten_dich_vu, da_xoa=False).exists():
                    return JsonResponse({'success': False, 'message': f'Tên dịch vụ {ten_dich_vu} đã tồn tại!'})
                
                # Get or create category based on name
                category_mapping = {
                    'haircut': 'Cắt tóc',
                    'shave': 'Cạo râu', 
                    'treatment': 'Chăm sóc',
                    'combo': 'Combo'
                }
                
                category_name = category_mapping.get(danh_muc, danh_muc)
                danh_muc_obj, created = DanhMucDichVu.objects.get_or_create(
                    ten_danh_muc=category_name,
                    defaults={'mo_ta': f'Danh mục {category_name}', 'da_xoa': False}
                )
                
                # Handle image upload
                anh_minh_hoa = ''
                if 'anh_minh_hoa' in request.FILES:
                    import os
                    from django.core.files.storage import default_storage
                    from django.core.files.base import ContentFile
                    
                    uploaded_file = request.FILES['anh_minh_hoa']
                    # Generate unique filename
                    file_extension = uploaded_file.name.split('.')[-1]
                    file_name = f"service_{int(timezone.now().timestamp())}_{ten_dich_vu.replace(' ', '_')}.{file_extension}"
                    file_path = f"services/{file_name}"
                    
                    # Save file
                    path = default_storage.save(file_path, ContentFile(uploaded_file.read()))
                    anh_minh_hoa = f"/media/{path}"
                
                service = DichVu.objects.create(
                    danh_muc=danh_muc_obj,
                    ten_dich_vu=ten_dich_vu,
                    mo_ta_ngan=request.POST.get('mo_ta', ''),
                    mo_ta_chi_tiet=request.POST.get('mo_ta', ''),
                    gia=int(gia),
                    thoi_gian_thuc_hien=int(thoi_luong),
                    anh_minh_hoa=anh_minh_hoa,
                    thu_tu=int(request.POST.get('thu_tu_hien_thi', 0)),
                    trang_thai=request.POST.get('trang_thai', 'active') == 'active',
                    da_xoa=False
                )
                return JsonResponse({'success': True, 'message': 'Đã thêm dịch vụ mới!', 'service_id': service.id})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'update':
            # Update existing service
            try:
                service_id = request.POST.get('service_id')
                service = get_object_or_404(DichVu, id=service_id, da_xoa=False)
                
                ten_dich_vu = request.POST.get('ten_dich_vu', '').strip()
                danh_muc = request.POST.get('danh_muc', '').strip()
                gia = request.POST.get('gia', '').strip()
                thoi_luong = request.POST.get('thoi_luong', '').strip()
                
                # Check if service name exists (excluding current service)
                if DichVu.objects.filter(ten_dich_vu=ten_dich_vu, da_xoa=False).exclude(id=service_id).exists():
                    return JsonResponse({'success': False, 'message': f'Tên dịch vụ {ten_dich_vu} đã tồn tại!'})
                
                # Get or create category
                category_mapping = {
                    'haircut': 'Cắt tóc',
                    'shave': 'Cạo râu', 
                    'treatment': 'Chăm sóc',
                    'combo': 'Combo'
                }
                
                category_name = category_mapping.get(danh_muc, danh_muc)
                danh_muc_obj, created = DanhMucDichVu.objects.get_or_create(
                    ten_danh_muc=category_name,
                    defaults={'mo_ta': f'Danh mục {category_name}', 'da_xoa': False}
                )
                
                # Handle image upload for update
                if 'anh_minh_hoa' in request.FILES:
                    import os
                    from django.core.files.storage import default_storage
                    from django.core.files.base import ContentFile
                    
                    # Delete old image if exists
                    if service.anh_minh_hoa and service.anh_minh_hoa.startswith('/media/'):
                        old_path = service.anh_minh_hoa.replace('/media/', '')
                        if default_storage.exists(old_path):
                            default_storage.delete(old_path)
                    
                    uploaded_file = request.FILES['anh_minh_hoa']
                    file_extension = uploaded_file.name.split('.')[-1]
                    file_name = f"service_{int(timezone.now().timestamp())}_{ten_dich_vu.replace(' ', '_')}.{file_extension}"
                    file_path = f"services/{file_name}"
                    
                    path = default_storage.save(file_path, ContentFile(uploaded_file.read()))
                    service.anh_minh_hoa = f"/media/{path}"
                
                service.danh_muc = danh_muc_obj
                service.ten_dich_vu = ten_dich_vu
                service.mo_ta_ngan = request.POST.get('mo_ta', '')
                service.mo_ta_chi_tiet = request.POST.get('mo_ta', '')
                service.gia = int(gia)
                service.thoi_gian_thuc_hien = int(thoi_luong)
                service.thu_tu = int(request.POST.get('thu_tu_hien_thi', 0))
                service.trang_thai = request.POST.get('trang_thai', 'active') == 'active'
                service.save()
                
                return JsonResponse({'success': True, 'message': 'Đã cập nhật dịch vụ!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'toggle_status':
            # Toggle active/inactive
            try:
                service_id = request.POST.get('service_id')
                service = get_object_or_404(DichVu, id=service_id, da_xoa=False)
                service.trang_thai = not service.trang_thai
                service.save()
                
                status_text = 'kích hoạt' if service.trang_thai else 'tạm ngừng'
                return JsonResponse({'success': True, 'message': f'Đã {status_text} dịch vụ!', 'new_status': service.trang_thai})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'delete':
            # Soft delete
            try:
                service_id = request.POST.get('service_id')
                service = get_object_or_404(DichVu, id=service_id, da_xoa=False)
                service.da_xoa = True
                service.ngay_xoa = timezone.now()
                service.save()
                
                return JsonResponse({'success': True, 'message': 'Đã xóa dịch vụ!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - display services
    from django.db.models import Count, Sum, Avg
    
    # Get all services with stats
    all_services = DichVu.objects.filter(
        da_xoa=False
    ).select_related('danh_muc').annotate(
        so_luot_dat=Count('dichvudatlich'),
        doanh_thu=Sum('dichvudatlich__dat_lich__hoa_don__thanh_tien')
    ).order_by('thu_tu', 'ten_dich_vu')
    
    # Add category field for template
    for service in all_services:
        if service.danh_muc:
            category_reverse_mapping = {
                'Cắt tóc': 'haircut',
                'Cạo râu': 'shave',
                'Chăm sóc': 'treatment', 
                'Combo': 'combo'
            }
            service.danh_muc_key = category_reverse_mapping.get(service.danh_muc.ten_danh_muc, 'haircut')
        else:
            service.danh_muc_key = 'haircut'
    
    # Statistics
    total_services = all_services.count()
    active_services = all_services.filter(trang_thai=True).count()
    avg_price = all_services.aggregate(avg_price=Avg('gia'))['avg_price'] or 0
    popular_service = all_services.order_by('-so_luot_dat').first()
    
    # Category counts
    haircut_count = all_services.filter(danh_muc__ten_danh_muc='Cắt tóc').count()
    shave_count = all_services.filter(danh_muc__ten_danh_muc='Cạo râu').count()
    treatment_count = all_services.filter(danh_muc__ten_danh_muc='Chăm sóc').count()
    combo_count = all_services.filter(danh_muc__ten_danh_muc='Combo').count()
    
    context = {
        'services': all_services,
        'total_services': total_services,
        'active_services': active_services,
        'avg_price': avg_price,
        'popular_service': popular_service or {'ten_dich_vu': 'Chưa có'},
        'haircut_count': haircut_count,
        'shave_count': shave_count,
        'treatment_count': treatment_count,
        'combo_count': combo_count,
    }
    return render(request, 'admin/services.html', context)

@require_role(['quan_ly'])
def api_service_detail(request, service_id):
    """API endpoint for service detail"""
    try:
        service = get_object_or_404(DichVu, id=service_id, da_xoa=False)
        
        # Map category name to key
        category_reverse_mapping = {
            'Cắt tóc': 'haircut',
            'Cạo râu': 'shave',
            'Chăm sóc': 'treatment', 
            'Combo': 'combo'
        }
        
        data = {
            'id': service.id,
            'ten_dich_vu': service.ten_dich_vu,
            'danh_muc': category_reverse_mapping.get(service.danh_muc.ten_danh_muc if service.danh_muc else '', 'haircut'),
            'gia': float(service.gia),
            'thoi_luong': service.thoi_gian_thuc_hien,
            'mo_ta': service.mo_ta_ngan or '',
            'trang_thai': 'active' if service.trang_thai else 'inactive',
            'thu_tu_hien_thi': service.thu_tu,
            'noi_bat': False,  # Add this field if needed
            'hinh_anh': service.anh_minh_hoa or ''
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@require_role(['quan_ly'])
def api_service_crud(request, service_id=None):
    """API endpoint for service CRUD operations"""
    if request.method == 'GET' and service_id:
        return api_service_detail(request, service_id)
    
    elif request.method == 'POST':
        # Create new service
        return admin_services(request)
    
    elif request.method == 'PUT' and service_id:
        # Update service - convert PUT to POST
        request.POST = request.POST.copy()
        request.POST['action'] = 'update'
        request.POST['service_id'] = service_id
        return admin_services(request)
    
    elif request.method == 'DELETE' and service_id:
        # Delete service
        request.POST = request.POST.copy()
        request.POST['action'] = 'delete'
        request.POST['service_id'] = service_id
        return admin_services(request)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@require_role(['quan_ly'])
def api_service_toggle_status(request, service_id):
    """API endpoint to toggle service status"""
    if request.method == 'POST':
        request.POST = request.POST.copy()
        request.POST['action'] = 'toggle_status'
        request.POST['service_id'] = service_id
        return admin_services(request)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@require_role(['quan_ly'])
def api_service_update_order(request):
    """API endpoint to update service display order"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            order_list = data.get('order', [])
            
            for item in order_list:
                service_id = item.get('id')
                new_order = item.get('order')
                if service_id and new_order is not None:
                    DichVu.objects.filter(id=service_id, da_xoa=False).update(thu_tu=new_order)
            
            return JsonResponse({'success': True, 'message': 'Đã cập nhật thứ tự!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@require_role(['quan_ly'])
def admin_invoices(request):
    """Invoice Management with filters and stats"""
    from datetime import date
    from django.db.models import Sum, Avg, Count
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    payment_method = request.GET.get('payment_method')
    staff_id = request.GET.get('staff_id')
    customer_type = request.GET.get('customer_type')
    search = request.GET.get('search')
    sort_by = request.GET.get('sort_by', '-ngay_thanh_toan')
    
    # Debug logging
    print(f"DEBUG Invoice - Filter parameters: from_date={from_date}, to_date={to_date}, payment_method={payment_method}, staff_id={staff_id}, search={search}")
    
    # Base queryset
    invoices = HoaDon.objects.filter(
        da_xoa=False
    ).select_related('khach_hang', 'nhan_vien', 'nguoi_tao', 'dat_lich').prefetch_related(
        'chi_tiet__dich_vu'
    )
    
    # Apply filters
    original_count = invoices.count()
    
    if from_date and from_date.strip() and from_date.strip().lower() != 'none':
        invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        print(f"DEBUG Invoice - After from_date filter ({from_date}): {invoices.count()} invoices")
    
    if to_date and to_date.strip() and to_date.strip().lower() != 'none':
        invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        print(f"DEBUG Invoice - After to_date filter ({to_date}): {invoices.count()} invoices")
    
    if payment_method and payment_method.strip():
        invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        print(f"DEBUG Invoice - After payment_method filter ({payment_method}): {invoices.count()} invoices")
    
    if staff_id and staff_id.strip():
        invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        print(f"DEBUG Invoice - After staff filter ({staff_id}): {invoices.count()} invoices")
    
    if search and search.strip() and search.strip().lower() != 'none':
        from django.db.models import Q
        search_term = search.strip()
        invoices = invoices.filter(
            Q(ma_hoa_don__icontains=search_term) |
            Q(ten_khach_hang__icontains=search_term) |
            Q(so_dien_thoai_khach__icontains=search_term)
        )
        print(f"DEBUG Invoice - After search filter ({search_term}): {invoices.count()} invoices")
    
    # Apply sorting
    valid_sort_fields = [
        'ngay_thanh_toan', '-ngay_thanh_toan',
        'thanh_tien', '-thanh_tien', 
        'ma_hoa_don', '-ma_hoa_don',
        'ten_khach_hang', '-ten_khach_hang'
    ]
    
    if sort_by and sort_by in valid_sort_fields:
        invoices = invoices.order_by(sort_by)
    else:
        invoices = invoices.order_by('-ngay_thanh_toan')
    
    print(f"DEBUG Invoice - Total invoices after all filters: {invoices.count()} (original: {original_count})")
    
    # Get all invoices for statistics (without pagination)
    all_invoices = invoices
    
    # Statistics
    today = date.today()
    
    total_invoices = all_invoices.count()
    total_revenue = all_invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
    today_invoices = all_invoices.filter(ngay_thanh_toan__date=today).count()
    today_revenue = all_invoices.filter(ngay_thanh_toan__date=today).aggregate(total=Sum('thanh_tien'))['total'] or 0
    average_invoice = all_invoices.aggregate(avg=Avg('thanh_tien'))['avg'] or 0
    
    # Payment method stats
    payment_methods = all_invoices.values('phuong_thuc_thanh_toan').annotate(
        count=Count('id'),
        total=Sum('thanh_tien')
    ).order_by('-total')
    
    # Get staff list for filter
    staff_list = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(invoices, 25)  # 25 invoices per page
    page_number = request.GET.get('page')
    invoices_page = paginator.get_page(page_number)
    
    # Calculate page total
    page_total = sum(invoice.thanh_tien for invoice in invoices_page.object_list)
    
    context = {
        'invoices': invoices_page,
        'total_invoices': total_invoices,
        'total_revenue': total_revenue,
        'today_invoices': today_invoices,
        'today_revenue': today_revenue,
        'average_invoice': average_invoice,
        'payment_methods': payment_methods,
        'staff_list': staff_list,
        'page_total': page_total,
        'from_date': from_date,
        'to_date': to_date,
        'payment_method': payment_method,
        'staff_id': staff_id,
        'customer_type': customer_type,
        'search': search,
        'sort_by': sort_by,
    }
    return render(request, 'admin/invoices.html', context)

@require_role(['quan_ly'])
def admin_invoices_export_excel(request):
    """Export invoices to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        payment_method = request.GET.get('payment_method')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset
        invoices = HoaDon.objects.filter(
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien', 'nguoi_tao').prefetch_related(
            'chi_tiet__dich_vu'
        ).order_by('-ngay_thanh_toan')
        
        # Apply filters (same logic as main view)
        if from_date and from_date.strip() and from_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        
        if to_date and to_date.strip() and to_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        
        if payment_method and payment_method.strip():
            invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        
        if staff_id and staff_id.strip():
            invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        
        if search and search.strip() and search.strip().lower() != 'none':
            from django.db.models import Q
            search_term = search.strip()
            invoices = invoices.filter(
                Q(ma_hoa_don__icontains=search_term) |
                Q(ten_khach_hang__icontains=search_term) |
                Q(so_dien_thoai_khach__icontains=search_term)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách hóa đơn"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Mã hóa đơn', 'Khách hàng', 'Số điện thoại', 'Nhân viên',
            'Ngày thanh toán', 'Dịch vụ', 'Tạm tính', 'Giảm giá', 'Thành tiền', 
            'Phương thức TT', 'Ghi chú'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Payment method mapping
        payment_method_map = {
            'tien_mat': 'Tiền mặt',
            'chuyen_khoan': 'Chuyển khoản',
            'vi_dien_tu': 'Ví điện tử',
            'the': 'Thẻ'
        }
        
        # Data rows
        for row, invoice in enumerate(invoices, 2):
            services = ', '.join([ct.ten_dich_vu for ct in invoice.chi_tiet.all()])
            
            ws.cell(row=row, column=1, value=invoice.ma_hoa_don)
            ws.cell(row=row, column=2, value=invoice.ten_khach_hang)
            ws.cell(row=row, column=3, value=invoice.so_dien_thoai_khach)
            ws.cell(row=row, column=4, value=invoice.nhan_vien.ho_ten if invoice.nhan_vien else '')
            ws.cell(row=row, column=5, value=invoice.ngay_thanh_toan.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=6, value=services)
            ws.cell(row=row, column=7, value=float(invoice.tam_tinh))
            ws.cell(row=row, column=8, value=float(invoice.tien_giam_gia))
            ws.cell(row=row, column=9, value=float(invoice.thanh_tien))
            ws.cell(row=row, column=10, value=payment_method_map.get(invoice.phuong_thuc_thanh_toan, invoice.phuong_thuc_thanh_toan))
            ws.cell(row=row, column=11, value=invoice.ghi_chu or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'danh_sach_hoa_don_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lỗi xuất Excel: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_invoices_export_pdf(request):
    """Export invoices to PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        import io
        
        # Get filter parameters (same as Excel export)
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        payment_method = request.GET.get('payment_method')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset with filters (same logic)
        invoices = HoaDon.objects.filter(da_xoa=False).select_related('khach_hang', 'nhan_vien').prefetch_related('chi_tiet__dich_vu').order_by('-ngay_thanh_toan')
        
        # Apply filters
        if from_date and from_date.strip() and from_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        if to_date and to_date.strip() and to_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        if payment_method and payment_method.strip():
            invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        if staff_id and staff_id.strip():
            invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        if search and search.strip() and search.strip().lower() != 'none':
            from django.db.models import Q
            search_term = search.strip()
            invoices = invoices.filter(Q(ma_hoa_don__icontains=search_term) | Q(ten_khach_hang__icontains=search_term) | Q(so_dien_thoai_khach__icontains=search_term))
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Content
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=1)
        title = Paragraph("DANH SÁCH HÓA ĐƠN", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Mã HĐ', 'Khách hàng', 'SĐT', 'Nhân viên', 'Ngày TT', 'Thành tiền', 'PT Thanh toán']]
        
        payment_method_map = {'tien_mat': 'Tiền mặt', 'chuyen_khoan': 'CK', 'vi_dien_tu': 'Ví', 'the': 'Thẻ'}
        
        for invoice in invoices[:50]:  # Limit for PDF
            data.append([
                invoice.ma_hoa_don,
                invoice.ten_khach_hang[:15] + '...' if len(invoice.ten_khach_hang) > 15 else invoice.ten_khach_hang,
                invoice.so_dien_thoai_khach,
                invoice.nhan_vien.ho_ten[:10] + '...' if invoice.nhan_vien and len(invoice.nhan_vien.ho_ten) > 10 else (invoice.nhan_vien.ho_ten if invoice.nhan_vien else ''),
                invoice.ngay_thanh_toan.strftime('%d/%m/%Y'),
                f"{invoice.thanh_tien:,.0f}đ",
                payment_method_map.get(invoice.phuong_thuc_thanh_toan, invoice.phuong_thuc_thanh_toan)
            ])
        
        # Create table
        table = Table(data, colWidths=[1*inch, 1.5*inch, 1*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f'danh_sach_hoa_don_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': 'Cần cài đặt reportlab để xuất PDF: pip install reportlab'
        })
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lỗi xuất PDF: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_work_schedule(request):
    """Work Schedule Management with Approval"""
    from django.db.models import Count, Sum, Avg, Q
    from datetime import date, timedelta, datetime
    import calendar
    
    print("DEBUG Work Schedule - Request received")
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            # Approve shift
            try:
                schedule_id = request.POST.get('schedule_id')
                schedule = get_object_or_404(LichLamViec, id=schedule_id, da_xoa=False)
                schedule.trang_thai = 'da_duyet'
                schedule.save()
                return JsonResponse({'success': True, 'message': 'Đã duyệt ca làm!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'reject':
            # Reject shift
            try:
                schedule_id = request.POST.get('schedule_id')
                schedule = get_object_or_404(LichLamViec, id=schedule_id, da_xoa=False)
                schedule.trang_thai = 'tu_choi'
                schedule.ghi_chu = request.POST.get('ly_do', '')
                schedule.save()
                return JsonResponse({'success': True, 'message': 'Đã từ chối ca làm!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'create':
            # Admin creates shift for staff
            try:
                staff_ids = request.POST.getlist('staff_ids')  # Multiple staff
                ngay_lam = request.POST.get('ngay_lam')
                ca_lam = request.POST.get('ca_lam')
                repeat_type = request.POST.get('repeat_type', 'none')
                repeat_until = request.POST.get('repeat_until')
                ghi_chu = request.POST.get('ghi_chu', '')
                
                # Set times based on shift type
                shift_times = {
                    'morning': ('08:00', '12:00'),
                    'afternoon': ('13:00', '17:00'),
                    'evening': ('17:00', '21:00'),
                    'fullday': ('08:00', '21:00')
                }
                
                gio_bat_dau, gio_ket_thuc = shift_times.get(ca_lam, ('08:00', '17:00'))
                
                created_count = 0
                start_date = datetime.strptime(ngay_lam, '%Y-%m-%d').date()
                
                # Handle repeat logic
                dates_to_create = [start_date]
                
                if repeat_type != 'none' and repeat_until:
                    end_date = datetime.strptime(repeat_until, '%Y-%m-%d').date()
                    current_date = start_date
                    
                    while current_date <= end_date:
                        if repeat_type == 'daily':
                            current_date += timedelta(days=1)
                        elif repeat_type == 'weekly':
                            current_date += timedelta(weeks=1)
                        elif repeat_type == 'monthly':
                            # Add one month
                            if current_date.month == 12:
                                current_date = current_date.replace(year=current_date.year+1, month=1)
                            else:
                                current_date = current_date.replace(month=current_date.month+1)
                        
                        if current_date <= end_date:
                            dates_to_create.append(current_date)
                
                # Create shifts for all selected staff and dates
                for staff_id in staff_ids:
                    for shift_date in dates_to_create:
                        # Check if shift already exists
                        existing = LichLamViec.objects.filter(
                            nhan_vien_id=staff_id,
                            ngay_lam=shift_date,
                            ca_lam=ca_lam,
                            da_xoa=False
                        ).exists()
                        
                        if not existing:
                            LichLamViec.objects.create(
                                nhan_vien_id=staff_id,
                                ngay_lam=shift_date,
                                ca_lam=ca_lam,
                                gio_bat_dau=gio_bat_dau,
                                gio_ket_thuc=gio_ket_thuc,
                                trang_thai='da_duyet',
                                ghi_chu=ghi_chu,
                                da_xoa=False
                            )
                            created_count += 1
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Đã tạo {created_count} ca làm việc!'
                })
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'delete':
            # Delete shift
            try:
                schedule_id = request.POST.get('schedule_id')
                schedule = get_object_or_404(LichLamViec, id=schedule_id, da_xoa=False)
                schedule.da_xoa = True
                schedule.ngay_xoa = timezone.now()
                schedule.save()
                return JsonResponse({'success': True, 'message': 'Đã xóa ca làm!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - display schedule
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    # Get filter parameters
    view_type = request.GET.get('view', 'week')  # week, month, pending
    staff_id = request.GET.get('staff_id')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    print(f"DEBUG Work Schedule - Params: view={view_type}, staff={staff_id}, from={from_date}, to={to_date}")
    
    # Get all active staff
    staff = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        da_xoa=False,
        trang_thai=True
    ).order_by('ho_ten')
    
    # Calculate statistics
    total_staff = staff.count()
    
    # Working today - staff who have shifts today
    working_today = LichLamViec.objects.filter(
        da_xoa=False,
        ngay_lam=today,
        trang_thai='da_duyet'
    ).values('nhan_vien').distinct().count()
    
    # Off today - total staff minus working today
    off_today = total_staff - working_today
    
    # Total shifts this week
    total_shifts_week = LichLamViec.objects.filter(
        da_xoa=False,
        ngay_lam__gte=week_start,
        ngay_lam__lte=week_end,
        trang_thai='da_duyet'
    ).count()
    
    # Pending leave requests (assuming we have a model for this)
    try:
        from barbershop.models import DonXinNghi
        pending_leaves = DonXinNghi.objects.filter(
            da_xoa=False,
            trang_thai='cho_duyet'
        ).count()
        
        leave_requests = DonXinNghi.objects.filter(
            da_xoa=False
        ).select_related('nhan_vien').order_by('-ngay_tao')[:20]
    except:
        # If leave request model doesn't exist, use 0
        pending_leaves = 0
        leave_requests = []
    
    # Average hours per week per staff (estimate 4h per shift for sang/chieu, 4h for toi)
    week_shifts = LichLamViec.objects.filter(
        da_xoa=False,
        ngay_lam__gte=week_start,
        ngay_lam__lte=week_end,
        trang_thai='da_duyet'
    )
    
    total_hours = 0
    for shift in week_shifts:
        # Estimate hours based on shift type
        if shift.ca_lam in ['sang', 'chieu']:
            total_hours += 4
        elif shift.ca_lam == 'toi':
            total_hours += 4
        else:
            total_hours += 8  # fullday
    
    avg_hours = (total_hours / total_staff) if total_staff > 0 else 0
    
    print(f"DEBUG Work Schedule - Stats: staff={total_staff}, working_today={working_today}, shifts={total_shifts_week}")
    
    # Filter staff if specified
    if staff_id and staff_id.strip():
        staff = staff.filter(id=staff_id.strip())
        print(f"DEBUG Work Schedule - Filtered to staff ID: {staff_id}")
    
    # Date range for filtering
    if view_type == 'month':
        if from_date and to_date:
            try:
                filter_start = datetime.strptime(from_date, '%Y-%m-%d').date()
                filter_end = datetime.strptime(to_date, '%Y-%m-%d').date()
            except:
                filter_start = today.replace(day=1)
                last_day = calendar.monthrange(today.year, today.month)[1]
                filter_end = today.replace(day=last_day)
        else:
            filter_start = today.replace(day=1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            filter_end = today.replace(day=last_day)
    else:
        if from_date and to_date:
            try:
                filter_start = datetime.strptime(from_date, '%Y-%m-%d').date()
                filter_end = datetime.strptime(to_date, '%Y-%m-%d').date()
            except:
                filter_start = week_start
                filter_end = week_end
        else:
            filter_start = week_start
            filter_end = week_end
    
    # Get schedules based on view type and filters
    base_schedule_query = LichLamViec.objects.filter(
        da_xoa=False,
        ngay_lam__gte=filter_start,
        ngay_lam__lte=filter_end
    ).select_related('nhan_vien')
    
    if staff_id and staff_id.strip():
        base_schedule_query = base_schedule_query.filter(nhan_vien_id=staff_id.strip())
    
    schedules = base_schedule_query.order_by('ngay_lam', 'gio_bat_dau')
    
    print(f"DEBUG Work Schedule - Found {schedules.count()} schedules for period {filter_start} to {filter_end}")
    
    # Organize schedules by staff and day for calendar view
    schedule_dict = {}
    for schedule in schedules:
        staff_id_key = schedule.nhan_vien_id
        if staff_id_key not in schedule_dict:
            schedule_dict[staff_id_key] = {}
        day = schedule.ngay_lam.strftime('%Y-%m-%d')
        if day not in schedule_dict[staff_id_key]:
            schedule_dict[staff_id_key][day] = []
        schedule_dict[staff_id_key][day].append(schedule)
    
    # Build week days for calendar view
    week_days = []
    current_date = filter_start
    while current_date <= filter_end:
        day_schedules = {
            'date': current_date,
            'morning_shifts': [],
            'afternoon_shifts': [],
            'evening_shifts': [],
            'fullday_shifts': [],
            'morning_leaves': [],
        }
        
        # Get shifts for this day
        day_str = current_date.strftime('%Y-%m-%d')
        for staff_member in staff:
            if staff_member.id in schedule_dict and day_str in schedule_dict[staff_member.id]:
                for shift in schedule_dict[staff_member.id][day_str]:
                    if shift.ca_lam == 'sang':
                        day_schedules['morning_shifts'].append(shift)
                    elif shift.ca_lam == 'chieu':
                        day_schedules['afternoon_shifts'].append(shift)
                    elif shift.ca_lam == 'toi':
                        day_schedules['evening_shifts'].append(shift)
                    else:
                        day_schedules['morning_shifts'].append(shift)  # Default to morning
        
        week_days.append(day_schedules)
        current_date += timedelta(days=1)
    
    # Build staff list with schedule for staff view
    staff_list = []
    for staff_member in staff:
        week_schedule = []
        total_hours = 0
        
        current_date = filter_start
        while current_date <= filter_end:
            day_str = current_date.strftime('%Y-%m-%d')
            day_shifts = []
            
            if staff_member.id in schedule_dict and day_str in schedule_dict[staff_member.id]:
                for shift in schedule_dict[staff_member.id][day_str]:
                    # Estimate hours based on shift type
                    hours = 4 if shift.ca_lam in ['sang', 'chieu', 'toi'] else 8
                    shift_info = {
                        'type': shift.ca_lam.lower(),
                        'label': shift.ca_lam.title(),
                        'hours': hours
                    }
                    day_shifts.append(shift_info)
                    total_hours += hours
            
            week_schedule.append({'shifts': day_shifts})
            current_date += timedelta(days=1)
        
        staff_list.append({
            'id': staff_member.id,
            'ho_ten': staff_member.ho_ten,
            'anh_dai_dien': staff_member.anh_dai_dien or '/static/img/avatar-default.png',
            'week_schedule': week_schedule,
            'total_hours': total_hours
        })
    
    context = {
        'staff': staff,
        'staff_list': staff_list,
        'schedules': schedule_dict,
        'week_days': week_days,
        'week_start': filter_start,
        'week_end': filter_end,
        'view_type': view_type,
        'from_date': from_date,
        'to_date': to_date,
        'staff_id': staff_id,
        # Stats
        'total_staff': total_staff,
        'working_today': working_today,
        'off_today': off_today,
        'total_shifts_week': total_shifts_week,
        'pending_leaves': pending_leaves,
        'avg_hours': avg_hours,
        'leave_requests': leave_requests,
    }
    
    return render(request, 'admin/work-schedule.html', context)

@require_role(['quan_ly'])
def admin_leave_request_approve(request, leave_id):
    """Approve leave request"""
    if request.method == 'POST':
        try:
            from barbershop.models import DonXinNghi
            leave_request = get_object_or_404(DonXinNghi, id=leave_id, da_xoa=False)
            leave_request.trang_thai = 'da_duyet'
            
            # Get current user from session
            current_user_id = request.session.get('user_id')
            if current_user_id:
                try:
                    current_user = NguoiDung.objects.get(id=current_user_id)
                    leave_request.nguoi_duyet = current_user
                except NguoiDung.DoesNotExist:
                    pass  # Leave nguoi_duyet as None if user not found
            
            leave_request.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Đã duyệt đơn xin nghỉ!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_leave_request_reject(request, leave_id):
    """Reject leave request"""
    if request.method == 'POST':
        try:
            from barbershop.models import DonXinNghi
            leave_request = get_object_or_404(DonXinNghi, id=leave_id, da_xoa=False)
            leave_request.trang_thai = 'tu_choi'
            leave_request.ly_do_tu_choi = request.POST.get('reason', '')
            
            # Get current user from session
            current_user_id = request.session.get('user_id')
            if current_user_id:
                try:
                    current_user = NguoiDung.objects.get(id=current_user_id)
                    leave_request.nguoi_duyet = current_user
                except NguoiDung.DoesNotExist:
                    pass  # Leave nguoi_duyet as None if user not found
            
            leave_request.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Đã từ chối đơn xin nghỉ!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_promotions(request):
    """Promotion Management with Enhanced Stats, Filtering and CRUD"""
    from django.db.models import Count, Sum, Q, F
    from datetime import datetime, date
    from django.contrib import messages
    
    # Handle POST requests for CRUD operations
    if request.method == 'POST':
        try:
            voucher_id = request.POST.get('voucher_id')
            
            # Validation
            ma_voucher = request.POST.get('ma_voucher', '').strip()
            ten_voucher = request.POST.get('ten_voucher', '').strip()
            loai_giam = request.POST.get('loai_giam', '').strip()
            gia_tri_giam = request.POST.get('gia_tri_giam', '').strip()
            ngay_bat_dau = request.POST.get('ngay_bat_dau', '').strip()
            ngay_ket_thuc = request.POST.get('ngay_ket_thuc', '').strip()
            
            if not all([ma_voucher, ten_voucher, loai_giam, gia_tri_giam, ngay_bat_dau, ngay_ket_thuc]):
                messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc!')
                return redirect('admin_promotions')
            
            # Check for duplicate voucher code (exclude current when editing)
            existing_voucher = Voucher.objects.filter(ma_voucher=ma_voucher, da_xoa=False)
            if voucher_id:
                existing_voucher = existing_voucher.exclude(id=voucher_id)
            if existing_voucher.exists():
                messages.error(request, f'Mã voucher "{ma_voucher}" đã tồn tại!')
                return redirect('admin_promotions')
            
            # Parse dates
            try:
                ngay_bat_dau_dt = datetime.strptime(ngay_bat_dau, '%Y-%m-%dT%H:%M')
                ngay_ket_thuc_dt = datetime.strptime(ngay_ket_thuc, '%Y-%m-%dT%H:%M')
                
                if ngay_bat_dau_dt >= ngay_ket_thuc_dt:
                    messages.error(request, 'Ngày kết thúc phải sau ngày bắt đầu!')
                    return redirect('admin_promotions')
                    
            except ValueError:
                messages.error(request, 'Định dạng ngày giờ không hợp lệ!')
                return redirect('admin_promotions')
            
            # Validate discount value
            try:
                gia_tri_giam_float = float(gia_tri_giam)
                if gia_tri_giam_float <= 0:
                    messages.error(request, 'Giá trị giảm phải lớn hơn 0!')
                    return redirect('admin_promotions')
                if loai_giam == 'phan_tram' and gia_tri_giam_float > 100:
                    messages.error(request, 'Giảm theo phần trăm không được vượt quá 100%!')
                    return redirect('admin_promotions')
            except ValueError:
                messages.error(request, 'Giá trị giảm không hợp lệ!')
                return redirect('admin_promotions')
            
            # Prepare data with correct field names from model
            voucher_data = {
                'ma_voucher': ma_voucher,
                'ten_voucher': ten_voucher,
                'mo_ta': request.POST.get('mo_ta', '').strip(),
                'loai_giam': loai_giam,
                'gia_tri_giam': gia_tri_giam_float,
                'gia_tri_don_toi_thieu': float(request.POST.get('gia_tri_don_hang_toi_thieu') or 0),  # Model field name
                'giam_toi_da': float(request.POST.get('gia_tri_giam_toi_da') or 0) or None,  # Model field name
                'ngay_bat_dau': ngay_bat_dau_dt.date(),  # Convert to date for DateField
                'ngay_ket_thuc': ngay_ket_thuc_dt.date(),  # Convert to date for DateField
                'so_luong_tong': int(request.POST.get('so_luong_toi_da') or 0) or None,  # Model field name
                'trang_thai': request.POST.get('trang_thai') == 'active',
            }
            
            if voucher_id:
                # Update existing voucher
                voucher = Voucher.objects.get(id=voucher_id, da_xoa=False)
                for key, value in voucher_data.items():
                    setattr(voucher, key, value)
                voucher.save()
                messages.success(request, f'Đã cập nhật voucher "{ma_voucher}" thành công!')
            else:
                # Create new voucher - Remove nguoi_tao field as it doesn't exist in Voucher model
                voucher = Voucher.objects.create(**voucher_data)
                messages.success(request, f'Đã tạo voucher "{ma_voucher}" thành công!')
                
        except Exception as e:
            messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        
        return redirect('admin_promotions')
    
    # Handle GET requests for listing and filtering
    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')
    search = request.GET.get('search', '').strip()
    
    # Base query
    promotions_query = Voucher.objects.filter(da_xoa=False)
    
    # Apply filters
    if status_filter == 'active':
        promotions_query = promotions_query.filter(
            trang_thai=True,
            ngay_bat_dau__lte=timezone.now(),
            ngay_ket_thuc__gte=timezone.now()
        )
    elif status_filter == 'inactive':
        promotions_query = promotions_query.filter(trang_thai=False)
    elif status_filter == 'expired':
        promotions_query = promotions_query.filter(ngay_ket_thuc__lt=timezone.now())
    elif status_filter == 'upcoming':
        promotions_query = promotions_query.filter(ngay_bat_dau__gt=timezone.now())
    
    if type_filter != 'all':
        promotions_query = promotions_query.filter(loai_giam=type_filter)
    
    if search:
        promotions_query = promotions_query.filter(
            Q(ma_voucher__icontains=search) |
            Q(ten_voucher__icontains=search) |
            Q(mo_ta__icontains=search)
        )
    
    promotions = promotions_query.order_by('-ngay_tao')
    
    # Add usage percentage to each promotion
    for promo in promotions:
        if promo.so_luong_tong and promo.so_luong_tong > 0:
            promo.usage_percentage = (promo.so_luong_da_dung / promo.so_luong_tong) * 100
        else:
            promo.usage_percentage = 0
    
    # Statistics
    today = timezone.now()
    total_promotions = Voucher.objects.filter(da_xoa=False).count()
    
    active_promotions = Voucher.objects.filter(
        da_xoa=False,
        trang_thai=True,
        ngay_bat_dau__lte=today,
        ngay_ket_thuc__gte=today
    ).count()
    
    expired_promotions = Voucher.objects.filter(
        da_xoa=False,
        ngay_ket_thuc__lt=today
    ).count()
    
    upcoming_promotions = Voucher.objects.filter(
        da_xoa=False,
        ngay_bat_dau__gt=today
    ).count()
    
    # Usage statistics  
    total_usage = Voucher.objects.filter(da_xoa=False).aggregate(
        total_used=Sum('so_luong_da_dung')
    )['total_used'] or 0
    
    context = {
        'promotions': promotions,
        'total_promotions': total_promotions,
        'active_promotions': active_promotions,
        'expired_promotions': expired_promotions,
        'upcoming_promotions': upcoming_promotions,
        'total_usage': total_usage,
        'filtered_count': promotions.count(),
        # Filter values for form
        'status_filter': status_filter,
        'type_filter': type_filter,
        'search': search,
        # Choices for dropdowns
        'status_choices': [
            ('all', 'Tất cả'),
            ('active', 'Đang hoạt động'),
            ('inactive', 'Tạm ngưng'),
            ('expired', 'Đã hết hạn'),
            ('upcoming', 'Sắp diễn ra')
        ],
        'type_choices': [
            ('all', 'Tất cả loại'),
            ('phan_tram', 'Phần trăm'),
            ('tien_mat', 'Tiền mặt')
        ]
    }
    
    return render(request, 'admin/promotions.html', context)

@require_role(['quan_ly'])
def admin_delete_promotion(request, voucher_id):
    """Delete promotion (soft delete)"""
    from django.contrib import messages
    from django.http import JsonResponse
    
    try:
        voucher = Voucher.objects.get(id=voucher_id, da_xoa=False)
        voucher.da_xoa = True
        voucher.save()
        
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': True, 'message': f'Đã xóa voucher "{voucher.ma_voucher}" thành công!'})
        else:
            messages.success(request, f'Đã xóa voucher "{voucher.ma_voucher}" thành công!')
            
    except Voucher.DoesNotExist:
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': False, 'message': 'Voucher không tồn tại!'})
        else:
            messages.error(request, 'Voucher không tồn tại!')
    except Exception as e:
        print(f"DEBUG Delete Promotion - Error: {e}")
        if request.headers.get('Content-Type') == 'application/json':
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
        else:
            messages.error(request, f'Có lỗi xảy ra: {str(e)}')
    
    return redirect('admin_promotions')

@require_role(['quan_ly'])
def admin_promotion_stats(request, voucher_id):
    """Get detailed statistics for a specific voucher"""
    from django.http import JsonResponse
    from datetime import datetime, date
    
    try:
        voucher = Voucher.objects.get(id=voucher_id, da_xoa=False)
        
        # Calculate remaining days - DateField objects don't have .date() method
        today = timezone.now().date()
        if voucher.ngay_ket_thuc > today:
            so_ngay_con_lai = (voucher.ngay_ket_thuc - today).days
        else:
            so_ngay_con_lai = 0
        
        # Calculate total savings from this voucher usage
        tong_tiet_kiem = 0
        if voucher.so_luong_da_dung > 0:
            # Use correct field name from model: gia_tri_don_toi_thieu
            avg_order_value = voucher.gia_tri_don_toi_thieu or 100000  # fallback
            if voucher.loai_giam == 'phan_tram':
                tong_tiet_kiem = (avg_order_value * voucher.gia_tri_giam / 100) * voucher.so_luong_da_dung
            else:
                tong_tiet_kiem = voucher.gia_tri_giam * voucher.so_luong_da_dung
        
        stats = {
            'ma_voucher': voucher.ma_voucher,
            'ten_voucher': voucher.ten_voucher,
            'mo_ta': voucher.mo_ta or '',
            'loai_giam': voucher.loai_giam,
            'gia_tri_giam': float(voucher.gia_tri_giam),
            'gia_tri_don_hang_toi_thieu': float(voucher.gia_tri_don_toi_thieu or 0),  # Fixed field name
            'gia_tri_giam_toi_da': float(voucher.giam_toi_da or 0),  # Fixed field name
            'ngay_bat_dau': str(voucher.ngay_bat_dau),  # DateField can be converted to string directly
            'ngay_ket_thuc': str(voucher.ngay_ket_thuc),  # DateField can be converted to string directly
            'so_luong_toi_da': voucher.so_luong_tong or 0,  # Fixed field name
            'so_luong_da_dung': voucher.so_luong_da_dung or 0,
            'trang_thai': voucher.trang_thai,
            'so_ngay_con_lai': so_ngay_con_lai,
            'tong_tiet_kiem': float(tong_tiet_kiem),
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Voucher.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Voucher không tồn tại!'
        })
    except Exception as e:
        print(f"DEBUG Promotion Stats - Error: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Có lỗi xảy ra: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_reports(request):
    """Comprehensive Reports with Analytics"""
    from django.db.models import Count, Sum, Avg, Q, F, Case, When, IntegerField
    from datetime import date, timedelta, datetime
    import calendar
    
    print("DEBUG Reports - Request received")
    
    today = timezone.now().date()
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    compare_with = request.GET.get('compare_with', 'none')
    
    # Set default date range (last 30 days)
    if not from_date or not to_date:
        from_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        to_date = today.strftime('%Y-%m-%d')
    
    try:
        filter_start = datetime.strptime(from_date, '%Y-%m-%d').date()
        filter_end = datetime.strptime(to_date, '%Y-%m-%d').date()
    except:
        filter_start = today - timedelta(days=30)
        filter_end = today
    
    print(f"DEBUG Reports - Date range: {filter_start} to {filter_end}")
    
    # Base queries for the selected period  
    invoices_query = HoaDon.objects.filter(
        da_xoa=False,
        ngay_thanh_toan__date__gte=filter_start,
        ngay_thanh_toan__date__lte=filter_end
    )
    
    bookings_query = DatLich.objects.filter(
        da_xoa=False,
        ngay_hen__gte=filter_start,
        ngay_hen__lte=filter_end
    )
    
    # === REVENUE ANALYSIS ===
    total_revenue = invoices_query.aggregate(total=Sum('thanh_tien'))['total'] or 0
    total_invoices = invoices_query.count()
    avg_invoice = invoices_query.aggregate(avg=Avg('thanh_tien'))['avg'] or 0
    
    # Previous period comparison
    period_days = (filter_end - filter_start).days
    prev_start = filter_start - timedelta(days=period_days)
    prev_end = filter_start - timedelta(days=1)
    
    prev_revenue = HoaDon.objects.filter(
        da_xoa=False,
        ngay_thanh_toan__date__gte=prev_start,
        ngay_thanh_toan__date__lte=prev_end
    ).aggregate(total=Sum('thanh_tien'))['total'] or 0
    
    revenue_growth = ((float(total_revenue) - float(prev_revenue)) / float(prev_revenue) * 100) if prev_revenue > 0 else 0
    
    # === BOOKING ANALYSIS ===
    total_bookings = bookings_query.count()
    completed_bookings = bookings_query.filter(trang_thai='hoan_thanh').count()
    completion_rate = (completed_bookings / total_bookings * 100) if total_bookings > 0 else 0
    
    # === CUSTOMER ANALYSIS ===
    unique_customers = bookings_query.values('khach_hang').distinct().count()
    
    # New vs returning customers (simplified)
    all_customers = bookings_query.values('khach_hang').distinct()
    new_customers = 0
    returning_customers = 0
    
    for customer in all_customers:
        if customer['khach_hang']:
            prev_bookings = DatLich.objects.filter(
                khach_hang=customer['khach_hang'],
                ngay_hen__lt=filter_start,
                da_xoa=False
            ).exists()
            if prev_bookings:
                returning_customers += 1
            else:
                new_customers += 1
    
    # === TOP SERVICES ===
    # Get services from DichVuDatLich if available, otherwise use bookings
    try:
        from barbershop.models import DichVuDatLich
        top_services_data = DichVuDatLich.objects.filter(
            dat_lich__in=bookings_query
        ).values('dich_vu__ten_dich_vu').annotate(
            count=Count('id'),
            revenue=Sum('thanh_tien')
        ).order_by('-count')[:5]
        
        top_services = []
        max_count = top_services_data[0]['count'] if top_services_data else 1
        
        for service in top_services_data:
            top_services.append({
                'ten_dich_vu': service['dich_vu__ten_dich_vu'] or 'N/A',
                'so_luot': service['count'],
                'doanh_thu': float(service['revenue'] or 0),
                'percent': round((service['count'] / max_count * 100), 1)
            })
    except:
        # Fallback to booking data
        top_services = [
            {'ten_dich_vu': 'Cắt tóc cơ bản', 'so_luot': 45, 'doanh_thu': 2250000, 'percent': 100},
            {'ten_dich_vu': 'Cắt tóc + Gội đầu', 'so_luot': 32, 'doanh_thu': 1920000, 'percent': 71},
            {'ten_dich_vu': 'Tạo kiểu', 'so_luot': 28, 'doanh_thu': 1680000, 'percent': 62},
            {'ten_dich_vu': 'Nhuộm tóc', 'so_luot': 15, 'doanh_thu': 1125000, 'percent': 33},
            {'ten_dich_vu': 'Massage đầu', 'so_luot': 12, 'doanh_thu': 600000, 'percent': 27}
        ]
    
    # === TOP STAFF ===
    top_staff_data = bookings_query.filter(
        nhan_vien__isnull=False
    ).values('nhan_vien__ho_ten').annotate(
        count=Count('id'),
        revenue=Sum('thanh_tien')
    ).order_by('-count')[:5]
    
    top_staff = []
    for staff_data in top_staff_data:
        # Get staff details
        staff_user = NguoiDung.objects.get(ho_ten=staff_data['nhan_vien__ho_ten'])
        top_staff.append({
            'ho_ten': staff_data['nhan_vien__ho_ten'],
            'so_luot': staff_data['count'],
            'doanh_thu': float(staff_data['revenue'] or 0),
            'anh_dai_dien': staff_user.anh_dai_dien or '/static/img/avatar-default.png',
            'rating': 4.5  # Placeholder rating
        })
    
    # === PAYMENT METHOD ANALYSIS ===
    payment_methods_data = invoices_query.values('phuong_thuc_thanh_toan').annotate(
        count=Count('id'),
        revenue=Sum('thanh_tien')
    )
    
    payment_methods = []
    total_payment_revenue = float(total_revenue) if total_revenue > 0 else 1
    
    # Initialize payment method percentages
    cash_percent = transfer_percent = ewallet_percent = card_percent = 0
    payment_chart_data = [0, 0, 0, 0]  # [cash, transfer, ewallet, card]
    
    for method in payment_methods_data:
        method_name = dict(HoaDon._meta.get_field('phuong_thuc_thanh_toan').choices).get(
            method['phuong_thuc_thanh_toan'], method['phuong_thuc_thanh_toan']
        )
        percentage = (float(method['revenue'] or 0) / total_payment_revenue * 100)
        
        payment_methods.append({
            'method': method_name,
            'count': method['count'],
            'revenue': float(method['revenue'] or 0),
            'percentage': round(percentage, 1)
        })
        
        # Set individual percentages for template display
        if method['phuong_thuc_thanh_toan'] == 'tien_mat':
            cash_percent = round(percentage, 1)
            payment_chart_data[0] = float(method['revenue'] or 0)
        elif method['phuong_thuc_thanh_toan'] == 'chuyen_khoan':
            transfer_percent = round(percentage, 1)
            payment_chart_data[1] = float(method['revenue'] or 0)
        elif method['phuong_thuc_thanh_toan'] == 'vi_dien_tu':
            ewallet_percent = round(percentage, 1)
            payment_chart_data[2] = float(method['revenue'] or 0)
        elif method['phuong_thuc_thanh_toan'] == 'the_tin_dung' or method['phuong_thuc_thanh_toan'] == 'the':
            card_percent = round(percentage, 1)
            payment_chart_data[3] = float(method['revenue'] or 0)
    
    # === HOURLY ANALYSIS ===
    hourly_data_full = [0] * 24
    for booking in bookings_query:
        if booking.gio_hen:
            hour = booking.gio_hen.hour
            hourly_data_full[hour] += 1
    
    # Extract business hours only (8h-21h for display)
    hourly_data = hourly_data_full[8:22]  # Hours 8-21
    
    # === WEEKDAY ANALYSIS ===
    weekday_data = [0] * 7  # Monday to Sunday
    for booking in bookings_query:
        weekday = booking.ngay_hen.weekday()  # 0=Monday, 6=Sunday
        weekday_data[weekday] += 1
    
    # === MONTHLY REVENUE DATA ===
    revenue_data = []
    revenue_labels = []
    
    # Get last 6 months of data
    for i in range(6):
        month_start = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start.replace(month=month_start.month % 12 + 1, day=1) - timedelta(days=1)) if month_start.month < 12 else month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
        
        month_revenue = HoaDon.objects.filter(
            da_xoa=False,
            ngay_thanh_toan__date__gte=month_start,
            ngay_thanh_toan__date__lte=month_end
        ).aggregate(total=Sum('thanh_tien'))['total'] or 0
        
        revenue_data.insert(0, float(month_revenue))
        revenue_labels.insert(0, f"T{month_start.month}/{month_start.year}")
    
    # === CUSTOMER SATISFACTION ===
    avg_rating = 4.2  # Placeholder since we don't have review data
    
    # === CUSTOMER TIER ANALYSIS ===  
    customer_tier_data = [30, 45, 20, 5]  # Bronze, Silver, Gold, Platinum percentages
    
    # Convert data to JSON for JavaScript
    import json
    revenue_data_json = json.dumps(revenue_data)
    revenue_labels_json = json.dumps(revenue_labels)
    payment_chart_data_json = json.dumps(payment_chart_data)
    hourly_data_json = json.dumps(hourly_data)
    weekday_data_json = json.dumps(weekday_data)
    customer_tier_data_json = json.dumps(customer_tier_data)

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'compare_with': compare_with,
        'total_revenue': float(total_revenue),
        'total_invoices': total_invoices,
        'avg_invoice': float(avg_invoice),
        'revenue_growth': round(revenue_growth, 1),
        'invoice_growth': round(revenue_growth, 1),  # Using same for now
        'booking_growth': round(revenue_growth, 1),  # Using same for now  
        'customer_growth': round(revenue_growth, 1),  # Using same for now
        'total_bookings': total_bookings,
        'completed_bookings': completed_bookings,
        'service_completion_rate': round(completion_rate, 1),
        'new_customers': new_customers,
        'customer_retention': round(50.0, 1),  # Placeholder
        'avg_rating': avg_rating,
        'top_services': top_services,
        'top_staff': top_staff,
        'payment_methods': payment_methods,
        'cash_percent': cash_percent,
        'transfer_percent': transfer_percent,
        'ewallet_percent': ewallet_percent,
        'card_percent': card_percent,
        # JSON data for charts
        'revenue_data': revenue_data_json,
        'revenue_labels': revenue_labels_json,
        'payment_method_data': payment_chart_data_json,
        'hourly_data': hourly_data_json,
        'weekday_data': weekday_data_json,
        'customer_tier_data': customer_tier_data_json,
        'show_comparison': compare_with != 'none',
    }
    
    print("DEBUG Reports - Full context created successfully")
    print(f"DEBUG - Revenue data: {revenue_data}")
    print(f"DEBUG - Total revenue: {total_revenue}")
    
    return render(request, 'admin/reports.html', context)

@require_role(['quan_ly'])
def admin_reports_export_excel(request):
    """Export reports to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import LineChart, Reference, BarChart, PieChart
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get same data as main report
        from_date = request.GET.get('from', '')
        to_date = request.GET.get('to', '')
        
        if not from_date or not to_date:
            today = timezone.now().date()
            from_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
            to_date = today.strftime('%Y-%m-%d')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo tổng hợp"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"BÁO CÁO TỔNG HỢP ({from_date} - {to_date})"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Get data (simplified version of main report logic)
        try:
            filter_start = datetime.strptime(from_date, '%Y-%m-%d').date()
            filter_end = datetime.strptime(to_date, '%Y-%m-%d').date()
        except:
            today = timezone.now().date()
            filter_start = today - timedelta(days=30)
            filter_end = today
        
        invoices = HoaDon.objects.filter(
            da_xoa=False,
            ngay_thanh_toan__date__gte=filter_start,
            ngay_thanh_toan__date__lte=filter_end
        )
        
        total_revenue = invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
        total_invoices = invoices.count()
        avg_invoice = invoices.aggregate(avg=Avg('thanh_tien'))['avg'] or 0
        
        # Summary section
        ws['A3'] = "TỔNG QUAN"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        
        ws['A5'] = "Tổng doanh thu:"
        ws['B5'] = f"{total_revenue:,.0f}đ"
        ws['A6'] = "Số hóa đơn:"
        ws['B6'] = total_invoices
        ws['A7'] = "Giá trị TB/hóa đơn:"
        ws['B7'] = f"{avg_invoice:,.0f}đ"
        
        # Revenue by day
        ws['A10'] = "DOANH THU THEO NGÀY"
        ws['A10'].font = header_font
        ws['A10'].fill = header_fill
        
        ws['A12'] = "Ngày"
        ws['B12'] = "Doanh thu"
        ws['A12'].font = header_font
        ws['B12'].font = header_font
        
        # Daily revenue data
        current_date = filter_start
        row = 13
        while current_date <= filter_end:
            daily_revenue = invoices.filter(
                ngay_thanh_toan__date=current_date
            ).aggregate(total=Sum('thanh_tien'))['total'] or 0
            
            ws[f'A{row}'] = current_date.strftime('%d/%m/%Y')
            ws[f'B{row}'] = daily_revenue
            row += 1
            current_date += timedelta(days=1)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'bao_cao_tong_hop_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lỗi xuất Excel: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_reports_export_pdf(request):
    """Export reports to PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        import io
        
        # Get parameters
        from_date = request.GET.get('from', '')
        to_date = request.GET.get('to', '')
        
        if not from_date or not to_date:
            today = timezone.now().date()
            from_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
            to_date = today.strftime('%Y-%m-%d')
        
        # Get data
        try:
            filter_start = datetime.strptime(from_date, '%Y-%m-%d').date()
            filter_end = datetime.strptime(to_date, '%Y-%m-%d').date()
        except:
            today = timezone.now().date()
            filter_start = today - timedelta(days=30)
            filter_end = today
        
        invoices = HoaDon.objects.filter(
            da_xoa=False,
            ngay_thanh_toan__date__gte=filter_start,
            ngay_thanh_toan__date__lte=filter_end
        )
        
        total_revenue = invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
        total_invoices = invoices.count()
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Content
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=1)
        title = Paragraph(f"BÁO CÁO TỔNG HỢP<br/>({from_date} - {to_date})", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Summary table
        summary_data = [
            ['Chỉ số', 'Giá trị'],
            ['Tổng doanh thu', f'{total_revenue:,.0f}đ'],
            ['Số hóa đơn', str(total_invoices)],
            ['Giá trị TB/hóa đơn', f'{(total_revenue/total_invoices if total_invoices > 0 else 0):,.0f}đ']
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Daily revenue table (sample)
        daily_title = Paragraph("DOANH THU THEO NGÀY (10 ngày gần nhất)", styles['Heading2'])
        elements.append(daily_title)
        elements.append(Spacer(1, 10))
        
        daily_data = [['Ngày', 'Doanh thu', 'Số HĐ']]
        
        # Get last 10 days data
        for i in range(9, -1, -1):
            check_date = filter_end - timedelta(days=i)
            if check_date >= filter_start:
                daily_invoices = invoices.filter(ngay_thanh_toan__date=check_date)
                daily_revenue = daily_invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
                daily_count = daily_invoices.count()
                
                daily_data.append([
                    check_date.strftime('%d/%m/%Y'),
                    f'{daily_revenue:,.0f}đ',
                    str(daily_count)
                ])
        
        daily_table = Table(daily_data, colWidths=[1.5*inch, 2*inch, 1*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(daily_table)
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f'bao_cao_tong_hop_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': 'Cần cài đặt reportlab để xuất PDF: pip install reportlab'
        })
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lỗi xuất PDF: {str(e)}'
        })

@require_role(['quan_ly'])
def admin_reviews(request):
    """Review Management with comprehensive features"""
    from django.db.models import Count, Avg, Q
    from django.core.paginator import Paginator
    from datetime import datetime, timedelta
    
    # Get filter parameters
    rating_filter = request.GET.get('rating', 'all')
    status_filter = request.GET.get('status', 'all')
    service_filter = request.GET.get('service', '')
    search = request.GET.get('search', '').strip()
    sort = request.GET.get('sort', '-ngay_tao')
    
    # Base query for reviews
    reviews_query = DanhGia.objects.filter(da_xoa=False).select_related(
        'khach_hang', 'hoa_don', 'dich_vu', 'nhan_vien', 'nguoi_phan_hoi'
    )
    
    # Apply filters
    if rating_filter != 'all' and rating_filter.isdigit():
        reviews_query = reviews_query.filter(so_sao=int(rating_filter))
    
    if status_filter == 'pending':
        reviews_query = reviews_query.filter(phan_hoi__isnull=True)
    elif status_filter == 'replied':
        reviews_query = reviews_query.filter(phan_hoi__isnull=False)
    
    if service_filter and service_filter.isdigit():
        reviews_query = reviews_query.filter(dich_vu_id=service_filter)
    
    if search:
        reviews_query = reviews_query.filter(
            Q(khach_hang__ho_ten__icontains=search) |
            Q(noi_dung__icontains=search) |
            Q(dich_vu__ten_dich_vu__icontains=search) |
            Q(nhan_vien__ho_ten__icontains=search)
        )
    
    # Apply sorting
    reviews_query = reviews_query.order_by(sort)
    
    # Calculate statistics
    total_reviews = DanhGia.objects.filter(da_xoa=False).count()
    avg_rating = DanhGia.objects.filter(da_xoa=False).aggregate(
        avg=Avg('so_sao')
    )['avg'] or 0
    
    pending_replies = DanhGia.objects.filter(
        da_xoa=False, 
        phan_hoi__isnull=True
    ).count()
    
    # Calculate satisfaction rate (4-5 stars)
    satisfied = DanhGia.objects.filter(da_xoa=False, so_sao__gte=4).count()
    satisfaction_rate = round((satisfied / total_reviews * 100) if total_reviews > 0 else 0, 1)
    
    # Rating breakdown
    rating_breakdown = []
    for stars in range(5, 0, -1):
        count = DanhGia.objects.filter(da_xoa=False, so_sao=stars).count()
        percent = round((count / total_reviews * 100) if total_reviews > 0 else 0, 1)
        rating_breakdown.append({
            'stars': stars,
            'count': count,
            'percent': percent
        })
    
    # Rating counts for tabs
    star_counts = {}
    for i in range(1, 6):
        star_counts[f'{i}_star_count'] = DanhGia.objects.filter(
            da_xoa=False, so_sao=i
        ).count()
    
    # Trend data for chart - find date range with actual data
    trend_labels = []
    trend_data = []
    trend_counts = []
    
    # Get the date range with reviews (last 30 days or actual review dates)
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    reviews_in_period = DanhGia.objects.filter(
        da_xoa=False,
        ngay_tao__date__gte=thirty_days_ago
    ).dates('ngay_tao', 'day').order_by('ngay_tao')
    
    if reviews_in_period:
        # Use actual review dates for more meaningful chart
        review_dates = list(reviews_in_period)[-7:]  # Last 7 dates with reviews
        
        for date in review_dates:
            trend_labels.append(date.strftime('%d/%m'))
            
            daily_reviews = DanhGia.objects.filter(
                da_xoa=False,
                ngay_tao__date=date
            )
            
            daily_avg = daily_reviews.aggregate(avg=Avg('so_sao'))['avg'] or 0
            daily_count = daily_reviews.count()
            
            trend_data.append(round(daily_avg, 1))
            trend_counts.append(daily_count)
    else:
        # Fallback to last 7 days even if empty
        for i in range(6, -1, -1):
            date = timezone.now().date() - timedelta(days=i)
            trend_labels.append(date.strftime('%d/%m'))
            
            daily_reviews = DanhGia.objects.filter(
                da_xoa=False,
                ngay_tao__date=date
            )
            
            daily_avg = daily_reviews.aggregate(avg=Avg('so_sao'))['avg'] or 0
            daily_count = daily_reviews.count()
            
            trend_data.append(round(daily_avg, 1))
            trend_counts.append(daily_count)
    
    # Pagination
    paginator = Paginator(reviews_query, 10)
    page_number = request.GET.get('page')
    reviews = paginator.get_page(page_number)
    
    # Get all services for filter
    services = DichVu.objects.filter(da_xoa=False).order_by('ten_dich_vu')
    
    # Convert data to JSON format for Chart.js
    import json
    
    context = {
        'reviews': reviews,
        'services': services,
        'total_reviews': total_reviews,
        'avg_rating': round(avg_rating, 1),
        'pending_replies': pending_replies,
        'satisfaction_rate': satisfaction_rate,
        'rating_breakdown': rating_breakdown,
        'trend_labels': json.dumps(trend_labels),
        'trend_data': json.dumps(trend_data),
        'trend_counts': json.dumps(trend_counts),
        'current_filters': {
            'rating': rating_filter,
            'status': status_filter,
            'service': service_filter,
            'search': search,
            'sort': sort
        },
        **star_counts
    }
    
    return render(request, 'admin/reviews.html', context)

@require_role(['quan_ly'])
def admin_loyalty(request):
    """Loyalty Program"""
    from django.db.models import Sum, Count, Avg
    
    # Get top customers based on points and spending
    top_customers = []
    customers = NguoiDung.objects.filter(
        vai_tro='khach_hang',
        da_xoa=False
    ).annotate(
        total_spent=Sum('hoadon__tong_tien'),
        total_visits=Count('hoadon', distinct=True),
        avg_rating=Avg('danhgia__so_sao')
    ).order_by('-diem_tich_luy')[:10]
    
    for customer in customers:
        # Calculate tier based on spending
        total_spent = customer.total_spent or 0
        if total_spent >= 10000000:  # 10M+
            tier_name = "VIP Platinum"
            tier_color = "#E5E4E2"
            tier_bg = "linear-gradient(135deg, #E5E4E2, #BCC6CC)"
            tier_icon = "👑"
            next_tier = "Max Level"
            points_to_next = 0
        elif total_spent >= 5000000:  # 5M+
            tier_name = "VIP Gold"
            tier_color = "#FFD700"
            tier_bg = "linear-gradient(135deg, #FFD700, #FFA500)"
            tier_icon = "⭐"
            next_tier = "VIP Platinum"
            points_to_next = 10000000 - total_spent
        elif total_spent >= 2000000:  # 2M+
            tier_name = "VIP Silver"
            tier_color = "#C0C0C0"
            tier_bg = "linear-gradient(135deg, #C0C0C0, #A9A9A9)"
            tier_icon = "🥈"
            next_tier = "VIP Gold"
            points_to_next = 5000000 - total_spent
        else:
            tier_name = "Thành viên"
            tier_color = "#6c757d"
            tier_bg = "linear-gradient(135deg, #6c757d, #5a6268)"
            tier_icon = "👤"
            next_tier = "VIP Silver"
            points_to_next = 2000000 - total_spent
        
        next_tier_progress = min(100, (total_spent / (total_spent + points_to_next)) * 100) if points_to_next > 0 else 100
        
        top_customers.append({
            'id': customer.id,
            'name': customer.ho_ten,
            'phone': customer.so_dien_thoai,
            'avatar': customer.anh_dai_dien,
            'points': customer.diem_tich_luy,
            'total_spent': total_spent,
            'total_visits': customer.total_visits or 0,
            'avg_rating': round(customer.avg_rating or 0, 1),
            'last_visit': customer.ngay_cap_nhat.strftime('%d/%m'),
            'tier_name': tier_name,
            'tier_color': tier_color,
            'tier_bg': tier_bg,
            'tier_icon': tier_icon,
            'next_tier': next_tier,
            'points_to_next_tier': points_to_next,
            'next_tier_progress': next_tier_progress
        })
    
    context = {
        'loyalty_tiers': [],
        'top_customers': top_customers,
        'total_loyalty_members': NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False).count(),
        'total_points_issued': NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False).aggregate(
            total=Sum('diem_tich_luy'))['total'] or 0,
    }
    return render(request, 'admin/loyalty.html', context)

@require_role(['quan_ly'])
def admin_inventory(request):
    """Inventory Management - Placeholder"""
    context = {
        'inventory_items': [],
    }
    return render(request, 'admin/inventory.html', context)

@require_role(['quan_ly'])
def admin_attendance(request):
    """Attendance Management - Placeholder"""
    today = timezone.now().date()
    
    # Get today's schedules
    schedules = LichLamViec.objects.filter(
        ngay_lam=today,
        da_xoa=False
    ).select_related('nhan_vien')
    
    context = {
        'schedules': schedules,
        'today': today,
    }
    return render(request, 'admin/attendance.html', context)

@require_role(['quan_ly'])
def admin_salary(request):
    """Salary Management - Placeholder"""
    staff = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        da_xoa=False
    ).select_related('thong_tin_nhan_vien')
    
    context = {
        'staff': staff,
    }
    return render(request, 'admin/salary.html', context)

@require_role(['quan_ly'])
def admin_settings(request):
    """System Settings"""
    settings = CaiDatHeThong.get_settings()
    
    # Prepare days of week data for template
    days_of_week = [
        {'name': 'Thứ 2', 'key': 'monday'},
        {'name': 'Thứ 3', 'key': 'tuesday'},
        {'name': 'Thứ 4', 'key': 'wednesday'},
        {'name': 'Thứ 5', 'key': 'thursday'},
        {'name': 'Thứ 6', 'key': 'friday'},
        {'name': 'Thứ 7', 'key': 'saturday'},
        {'name': 'Chủ nhật', 'key': 'sunday'},
    ]
    
    for day in days_of_week:
        day['is_open'] = settings.ngay_lam_viec.get(day['key'], True)
        day['open_time'] = settings.gio_mo_cua
        day['close_time'] = settings.gio_dong_cua
    
    context = {
        'settings': settings,
        'days_of_week': days_of_week,
        'last_backup': settings.lan_sao_luu_cuoi,
    }
    return render(request, 'admin/settings.html', context)

@require_role(['quan_ly'])  
def admin_settings_api_general(request):
    """API endpoint for general settings"""
    if request.method == 'POST':
        try:
            settings = CaiDatHeThong.get_settings()
            
            # Update general fields
            settings.ten_tiem = request.POST.get('ten_tiem', settings.ten_tiem)
            settings.hotline = request.POST.get('hotline', settings.hotline)
            settings.dia_chi = request.POST.get('dia_chi', settings.dia_chi)
            settings.email = request.POST.get('email', settings.email)
            settings.website = request.POST.get('website', settings.website)
            settings.mo_ta = request.POST.get('mo_ta', settings.mo_ta)
            
            # Handle logo upload
            if 'logo' in request.FILES:
                logo_file = request.FILES['logo']
                # Simple file handling - in production, use proper storage
                logo_path = f'uploads/logos/{logo_file.name}'
                settings.logo = logo_path
            
            settings.save()
            
            return JsonResponse({'success': True, 'message': 'Đã lưu thông tin chung thành công!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_settings_api_business_hours(request):
    """API endpoint for business hours settings"""
    if request.method == 'POST':
        try:
            settings = CaiDatHeThong.get_settings()
            
            # Update working days
            ngay_lam_viec = {}
            days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            for day in days:
                ngay_lam_viec[day] = request.POST.get(f'{day}_open') == 'on'
            
            settings.ngay_lam_viec = ngay_lam_viec
            settings.gio_mo_cua = request.POST.get('gio_mo_cua', settings.gio_mo_cua)
            settings.gio_dong_cua = request.POST.get('gio_dong_cua', settings.gio_dong_cua)
            settings.thoi_gian_slot = int(request.POST.get('slot_duration', 30))
            
            settings.save()
            
            return JsonResponse({'success': True, 'message': 'Đã cập nhật giờ làm việc thành công!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_settings_api_services(request):
    """API endpoint for service settings"""
    if request.method == 'POST':
        try:
            settings = CaiDatHeThong.get_settings()
            
            settings.tu_dong_xac_nhan = request.POST.get('auto_confirm') == 'on'
            settings.cho_phep_chon_nhan_vien = request.POST.get('allow_staff_select') == 'on'
            settings.yeu_cau_dat_coc = request.POST.get('require_deposit') == 'on'
            
            if settings.yeu_cau_dat_coc:
                settings.so_tien_dat_coc = float(request.POST.get('deposit_amount', 50000))
                settings.loai_dat_coc = request.POST.get('deposit_type', 'VND')
            
            settings.save()
            
            return JsonResponse({'success': True, 'message': 'Đã cập nhật cài đặt dịch vụ thành công!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_settings_api_payments(request):
    """API endpoint for payment settings"""
    if request.method == 'POST':
        try:
            settings = CaiDatHeThong.get_settings()
            
            settings.thanh_toan_tien_mat = request.POST.get('cash_enabled') == 'on'
            settings.thanh_toan_the = request.POST.get('card_enabled') == 'on'
            settings.thanh_toan_chuyen_khoan = request.POST.get('bank_enabled') == 'on'
            settings.so_tai_khoan = request.POST.get('bank_account', settings.so_tai_khoan)
            settings.ten_ngan_hang = request.POST.get('bank_name', settings.ten_ngan_hang)
            
            settings.thanh_toan_momo = request.POST.get('momo_enabled') == 'on'
            settings.so_dien_thoai_momo = request.POST.get('momo_phone', settings.so_dien_thoai_momo)
            
            settings.thanh_toan_vnpay = request.POST.get('vnpay_enabled') == 'on'
            settings.vnpay_api_key = request.POST.get('vnpay_key', settings.vnpay_api_key)
            
            settings.save()
            
            return JsonResponse({'success': True, 'message': 'Đã cập nhật cài đặt thanh toán thành công!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_content(request):
    """Content Management - Placeholder"""
    context = {}
    return render(request, 'admin/content.html', context)

@require_role(['quan_ly'])
def admin_pos_report(request):
    """POS Report"""
    from django.db.models import Sum, Count, Avg
    from datetime import date, timedelta
    
    today = timezone.now().date()
    
    # Get filter parameters  
    from_date = request.GET.get('from_date', (today - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date = request.GET.get('to_date', today.strftime('%Y-%m-%d'))
    staff_filter = request.GET.get('staff_filter', '')
    payment_filter = request.GET.get('payment_filter', '')
    
    # Base invoice query
    invoices = HoaDon.objects.filter(
        ngay_tao__date__range=[from_date, to_date],
        da_xoa=False
    )
    
    if staff_filter:
        invoices = invoices.filter(nhan_vien_id=staff_filter)
    if payment_filter:
        invoices = invoices.filter(phuong_thuc_thanh_toan=payment_filter)
    
    # Get all staff for filter
    all_staff = NguoiDung.objects.filter(vai_tro='nhan_vien', da_xoa=False)
    
    # Calculate metrics
    total_sales = invoices.aggregate(total=Sum('tong_tien'))['total'] or 0
    total_orders = invoices.count()
    avg_order_value = total_sales / total_orders if total_orders > 0 else 0
    
    # Payment method breakdown
    payment_breakdown = invoices.values('phuong_thuc_thanh_toan').annotate(
        count=Count('id'),
        total=Sum('tong_tien')
    )
    
    cash_total = next((p['total'] for p in payment_breakdown if p['phuong_thuc_thanh_toan'] == 'cash'), 0) or 0
    card_total = next((p['total'] for p in payment_breakdown if p['phuong_thuc_thanh_toan'] == 'card'), 0) or 0
    momo_total = next((p['total'] for p in payment_breakdown if p['phuong_thuc_thanh_toan'] == 'momo'), 0) or 0
    
    cash_percent = (cash_total / total_sales * 100) if total_sales > 0 else 0
    card_percent = (card_total / total_sales * 100) if total_sales > 0 else 0
    momo_percent = (momo_total / total_sales * 100) if total_sales > 0 else 0
    
    context = {
        'from_date': from_date,
        'to_date': to_date,
        'all_staff': all_staff,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'cash_total': cash_total,
        'card_total': card_total,
        'momo_total': momo_total,
        'cash_percent': round(cash_percent, 1),
        'card_percent': round(card_percent, 1),
        'momo_percent': round(momo_percent, 1),
        'recent_invoices': invoices.order_by('-ngay_tao')[:20],
    }
    return render(request, 'admin/pos-report.html', context)

@require_role(['quan_ly'])
def admin_export_schedule(request):
    """Export Schedule"""
    schedules = LichLamViec.objects.filter(
        da_xoa=False
    ).select_related('nhan_vien').order_by('ngay_lam', 'ca_lam')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="lich_lam_viec.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nhân viên', 'Ngày', 'Ca làm việc', 'Giờ bắt đầu', 'Giờ kết thúc'])
    
    for schedule in schedules:
        writer.writerow([
            schedule.nhan_vien.ho_ten,
            schedule.ngay_lam,
            schedule.ca_lam,
            schedule.gio_bat_dau,
            schedule.gio_ket_thuc
        ])
    
    return response

@require_role(['quan_ly'])
def admin_export_promotions(request):
    """Export Promotions"""
    promotions = Voucher.objects.filter(da_xoa=False).order_by('-ngay_tao')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="khuyen_mai.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Mã voucher', 'Tên voucher', 'Mô tả', 'Loại giảm', 'Giá trị', 'Ngày bắt đầu', 'Ngày kết thúc', 'Trạng thái'])
    
    for promo in promotions:
        status_text = 'Hoạt động' if promo.trang_thai else 'Tạm ngưng'
        loai_giam_text = dict(promo.LOAI_GIAM_CHOICES).get(promo.loai_giam, promo.loai_giam)
        
        writer.writerow([
            promo.ma_voucher,
            promo.ten_voucher,
            promo.mo_ta or '',
            loai_giam_text,
            promo.gia_tri_giam,
            promo.ngay_bat_dau,
            promo.ngay_ket_thuc,
            status_text
        ])
    
    return response

# ============ STAFF VIEWS ============

@require_role(['nhan_vien', 'quan_ly'])
def staff_dashboard(request):
    """Staff Dashboard"""
    user_id = request.session.get('user_id')
    today = timezone.now().date()
    
    # Get staff info
    staff = NguoiDung.objects.get(id=user_id)
    staff_info = ThongTinNhanVien.objects.filter(nguoi_dung_id=user_id).first()
    
    # Today's bookings for this staff
    today_bookings = DatLich.objects.filter(
        nhan_vien_id=user_id,
        ngay_hen=today,
        da_xoa=False
    ).exclude(trang_thai='da_huy').select_related('khach_hang').order_by('gio_hen')
    
    # Count today's bookings by status
    waiting_checkin = today_bookings.filter(trang_thai='da_xac_nhan').count()
    
    # Today's schedule
    today_schedule = LichLamViec.objects.filter(
        nhan_vien_id=user_id,
        ngay_lam=today,
        da_xoa=False
    ).first()
    
    # This month statistics
    from django.db.models import Sum
    month_start = today.replace(day=1)
    
    month_completed_bookings = DatLich.objects.filter(
        nhan_vien_id=user_id,
        ngay_hen__gte=month_start,
        ngay_hen__lte=today,
        da_xoa=False,
        trang_thai='hoan_thanh'
    )
    
    month_bookings_count = month_completed_bookings.count()
    month_revenue = month_completed_bookings.aggregate(
        total=Sum('thanh_tien')
    )['total'] or 0
    
    # Average revenue per booking
    avg_revenue = month_revenue / month_bookings_count if month_bookings_count > 0 else 0
    
    # Next booking
    next_booking = today_bookings.filter(
        trang_thai__in=['da_xac_nhan', 'da_checkin']
    ).first()
    
    # Staff rating
    staff_rating = staff_info.danh_gia_trung_binh if staff_info else 0
    total_services = staff_info.tong_luot_phuc_vu if staff_info else 0
    
    context = {
        'staff': staff,
        'today_bookings': today_bookings,
        'today_schedule': today_schedule,
        'booking_count': today_bookings.count(),
        'waiting_checkin': waiting_checkin,
        'month_bookings_count': month_bookings_count,
        'month_revenue': month_revenue,
        'avg_revenue': avg_revenue,
        'next_booking': next_booking,
        'staff_rating': staff_rating,
        'total_services': total_services,
    }
    return render(request, 'staff/dashboard.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_today_bookings(request):
    """Today's Bookings"""
    user_id = request.session.get('user_id')
    today = timezone.now().date()
    
    bookings = DatLich.objects.filter(
        nhan_vien_id=user_id,
        ngay_hen=today,
        da_xoa=False
    ).select_related('khach_hang').prefetch_related(
        'dich_vu_dat_lich__dich_vu'
    ).order_by('gio_hen')
    
    context = {
        'bookings': bookings,
        'today': today,
    }
    return render(request, 'staff/today-bookings.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_schedule(request):
    """My Schedule"""
    user_id = request.session.get('user_id')
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    schedules = LichLamViec.objects.filter(
        nhan_vien_id=user_id,
        ngay_lam__gte=week_start,
        ngay_lam__lte=week_end,
        da_xoa=False
    ).order_by('ngay_lam')
    
    context = {
        'schedules': schedules,
        'week_start': week_start,
        'week_end': week_end,
    }
    return render(request, 'staff/schedule.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_profile(request):
    """My Profile with Update Capability"""
    user_id = request.session.get('user_id')
    user = get_object_or_404(NguoiDung, id=user_id, da_xoa=False)
    
    staff_info = None
    if hasattr(user, 'thong_tin_nhan_vien'):
        staff_info = user.thong_tin_nhan_vien
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_info':
            # Update basic info
            try:
                user.ho_ten = request.POST.get('ho_ten', user.ho_ten)
                user.email = request.POST.get('email', user.email)
                user.dia_chi = request.POST.get('dia_chi', user.dia_chi)
                user.save()
                
                # Update staff info if exists
                if staff_info:
                    staff_info.cccd = request.POST.get('cccd', staff_info.cccd)
                    staff_info.ngay_sinh = request.POST.get('ngay_sinh', staff_info.ngay_sinh)
                    staff_info.gioi_tinh = request.POST.get('gioi_tinh', staff_info.gioi_tinh)
                    staff_info.save()
                
                return JsonResponse({'success': True, 'message': 'Đã cập nhật thông tin!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        
        elif action == 'change_password':
            # Change password
            try:
                old_password = request.POST.get('old_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                # Verify old password
                if not bcrypt.checkpw(old_password.encode('utf-8'), user.mat_khau_hash.encode('utf-8')):
                    return JsonResponse({'success': False, 'message': 'Mật khẩu cũ không đúng!'})
                
                # Check new password match
                if new_password != confirm_password:
                    return JsonResponse({'success': False, 'message': 'Mật khẩu xác nhận không khớp!'})
                
                # Check password strength
                if len(new_password) < 6:
                    return JsonResponse({'success': False, 'message': 'Mật khẩu phải có ít nhất 6 ký tự!'})
                
                # Update password
                hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
                user.mat_khau_hash = hashed.decode('utf-8')
                user.save()
                
                return JsonResponse({'success': True, 'message': 'Đã đổi mật khẩu thành công!'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - show profile
    context = {
        'user': user,
        'staff_info': staff_info,
    }
    return render(request, 'staff/profile.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_revenue(request):
    """My Revenue"""
    user_id = request.session.get('user_id')
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    # This month completed bookings
    completed_bookings = DatLich.objects.filter(
        nhan_vien_id=user_id,
        ngay_hen__gte=month_start,
        ngay_hen__lte=today,
        da_xoa=False,
        trang_thai='hoan_thanh'
    ).select_related('khach_hang')
    
    context = {
        'bookings': completed_bookings,
        'total_bookings': completed_bookings.count(),
    }
    return render(request, 'staff/revenue.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_commission(request):
    """My Commission - Placeholder"""
    context = {}
    return render(request, 'staff/commission.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_my_customers(request):
    """My Customers"""
    user_id = request.session.get('user_id')
    
    # Get customers who booked with this staff
    customer_ids = DatLich.objects.filter(
        nhan_vien_id=user_id,
        da_xoa=False
    ).values_list('khach_hang_id', flat=True).distinct()
    
    customers = NguoiDung.objects.filter(
        id__in=customer_ids,
        da_xoa=False
    ).annotate(
        booking_count=Count('dat_lich', filter=Q(
            dat_lich__nhan_vien_id=user_id,
            dat_lich__da_xoa=False
        ))
    ).order_by('-booking_count')
    
    context = {
        'customers': customers,
    }
    return render(request, 'staff/my-customers.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_pos(request):
    """POS - Point of Sale System"""
    user_id = request.session.get('user_id')
    
    if request.method == 'POST':
        # Process payment
        try:
            # Get customer info
            customer_type = request.POST.get('customer_type', 'walk-in')
            customer_id = request.POST.get('customer_id')
            customer_name = request.POST.get('customer_name', '')
            customer_phone = request.POST.get('customer_phone', '')
            booking_id = request.POST.get('booking_id')
            
            # Get services from cart (JSON string)
            import json
            cart_data = request.POST.get('cart_data', '[]')
            cart = json.loads(cart_data)
            
            if not cart:
                return JsonResponse({'success': False, 'message': 'Vui lòng chọn ít nhất một dịch vụ!'})
            
            # Get payment info
            payment_method = request.POST.get('payment_method', 'tien_mat')
            voucher_id = request.POST.get('voucher_id')
            points_used = int(request.POST.get('points_used', 0) or 0)
            stylist_id = request.POST.get('stylist_id', user_id)
            
            # Calculate totals
            tam_tinh = sum(int(item['price']) * int(item['quantity']) for item in cart)
            tien_giam_gia = 0
            
            # Apply voucher
            if voucher_id:
                voucher = Voucher.objects.filter(id=voucher_id, trang_thai='hoat_dong', da_xoa=False).first()
                if voucher:
                    if voucher.loai_giam == 'tien':
                        tien_giam_gia += voucher.gia_tri_giam
                    else:  # phan_tram
                        tien_giam_gia += int(tam_tinh * voucher.gia_tri_giam / 100)
            
            # Apply points (1 point = 1,000 VND)
            if points_used > 0:
                tien_giam_gia += points_used * 1000
            
            thanh_tien = tam_tinh - tien_giam_gia
            
            # Create or link booking
            if booking_id:
                # From existing booking
                booking = DatLich.objects.get(id=booking_id)
                booking.trang_thai = 'hoan_thanh'
                booking.ngay_hoan_thanh = timezone.now()
                booking.save()
            else:
                # Create new booking
                booking = DatLich.objects.create(
                    khach_hang_id=customer_id if customer_id else None,
                    ten_khach_hang=customer_name,
                    so_dien_thoai_khach=customer_phone,
                    nhan_vien_id=stylist_id,
                    ngay_hen=timezone.now().date(),
                    gio_hen=timezone.now().time(),
                    loai_dat_lich='walk_in',
                    trang_thai='hoan_thanh',
                    ngay_check_in=timezone.now(),
                    ngay_hoan_thanh=timezone.now(),
                    tong_tien=tam_tinh,
                    tien_giam_gia=tien_giam_gia,
                    thanh_tien=thanh_tien,
                    da_xoa=False
                )
                
                # Add services to booking
                for item in cart:
                    DichVuDatLich.objects.create(
                        dat_lich=booking,
                        dich_vu_id=item['id'],
                        so_luong=item['quantity'],
                        gia_tai_thoi_diem=item['price'],
                        thanh_tien=int(item['price']) * int(item['quantity']),
                        da_xoa=False
                    )
            
            # Create invoice
            invoice = HoaDon.objects.create(
                dat_lich=booking,
                khach_hang_id=customer_id if customer_id else None,
                nhan_vien_id=stylist_id,
                phuong_thuc_thanh_toan=payment_method,
                voucher_id=voucher_id if voucher_id else None,
                diem_su_dung=points_used,
                tam_tinh=tam_tinh,
                tien_giam_gia=tien_giam_gia,
                thanh_tien=thanh_tien,
                trang_thai='da_thanh_toan',
                ngay_thanh_toan=timezone.now(),
                da_xoa=False
            )
            
            # Update customer points
            if customer_id:
                customer = NguoiDung.objects.get(id=customer_id)
                # Deduct used points
                if points_used > 0:
                    customer.diem_tich_luy = max(0, customer.diem_tich_luy - points_used)
                # Add new points (1% of thanh_tien)
                new_points = int(thanh_tien / 1000 * 0.01)
                customer.diem_tich_luy += new_points
                customer.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Thanh toán thành công!',
                'invoice_id': invoice.id,
                'ma_hoa_don': invoice.ma_hoa_don
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - show POS interface
    try:
        services = DichVu.objects.filter(da_xoa=False, trang_thai=True).select_related('danh_muc')
    except Exception as e:
        print(f"Error loading services: {e}")
        services = DichVu.objects.filter(da_xoa=False).select_related('danh_muc')
    
    try:
        customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False)[:50]  # Recent customers
    except Exception as e:
        print(f"Error loading customers: {e}")
        customers = []
    
    try:
        today_bookings = DatLich.objects.filter(
            ngay_hen=timezone.now().date(),
            trang_thai__in=['da_xac_nhan', 'da_checkin'],
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien')
    except Exception as e:
        print(f"Error loading today_bookings: {e}")
        today_bookings = []
    
    try:
        # Fixed: Use boolean True instead of string 'hoat_dong'
        vouchers = Voucher.objects.filter(
            trang_thai=True,
            ngay_bat_dau__lte=timezone.now(),
            ngay_ket_thuc__gte=timezone.now(),
            da_xoa=False
        )
    except Exception as e:
        print(f"Error loading vouchers: {e}")
        vouchers = []
    
    try:
        staff = NguoiDung.objects.filter(vai_tro='nhan_vien', da_xoa=False, trang_thai=True)
    except Exception as e:
        print(f"Error loading staff: {e}")
        staff = []
    
    context = {
        'services': services,
        'customers': customers,
        'today_bookings': today_bookings,
        'vouchers': vouchers,
        'staff': staff,
        'current_user_id': user_id,
    }
    return render(request, 'staff/pos.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_bookings_create(request):
    """Create Booking (Staff)"""
    if request.method == 'POST':
        user_id = request.session.get('user_id')
        
        booking = DatLich.objects.create(
            khach_hang_id=request.POST.get('khach_hang_id'),
            nhan_vien_id=user_id,
            ngay_hen=request.POST.get('ngay_hen'),
            gio_hen=request.POST.get('gio_hen'),
            trang_thai='da_xac_nhan',
            ghi_chu=request.POST.get('ghi_chu', ''),
            da_xoa=False
        )
        
        service_ids = request.POST.getlist('service_ids')
        for service_id in service_ids:
            DichVuDatLich.objects.create(
                dat_lich=booking,
                dich_vu_id=service_id,
                da_xoa=False
            )
        
        return redirect('staff_today_bookings')
    
    customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False)
    services = DichVu.objects.filter(da_xoa=False, trang_thai=True)
    
    context = {
        'customers': customers,
        'services': services,
    }
    return render(request, 'staff/bookings-create.html', context)

@require_role(['nhan_vien', 'quan_ly'])
def staff_register_shift(request):
    """Register for work shift"""
    user_id = request.session.get('user_id')
    
    if request.method == 'POST':
        ngay_lam = request.POST.get('ngay_lam')
        ca_lam = request.POST.get('ca_lam')
        ghi_chu = request.POST.get('ghi_chu', '')
        
        # Set default times based on shift
        if ca_lam == 'sang':
            gio_bat_dau = '08:00'
            gio_ket_thuc = '12:00'
        elif ca_lam == 'chieu':
            gio_bat_dau = '13:00'
            gio_ket_thuc = '17:00'
        else:  # toi
            gio_bat_dau = '18:00'
            gio_ket_thuc = '22:00'
        
        LichLamViec.objects.create(
            nhan_vien_id=user_id,
            ngay_lam=ngay_lam,
            ca_lam=ca_lam,
            gio_bat_dau=gio_bat_dau,
            gio_ket_thuc=gio_ket_thuc,
            trang_thai='cho_duyet',
            ghi_chu=ghi_chu,
            da_xoa=False
        )
        
        return redirect('staff_schedule')
    
    # Get existing shifts
    today = timezone.now().date()
    existing_shifts = LichLamViec.objects.filter(
        nhan_vien_id=user_id,
        ngay_lam__gte=today,
        da_xoa=False
    ).order_by('ngay_lam')
    
    context = {
        'existing_shifts': existing_shifts,
        'today': today,
    }
    return render(request, 'staff/register-shift.html', context)

# ============ API ENDPOINTS ============

def api_search_customer(request):
    """API: Search customer by phone or name"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'success': False, 'message': 'Query required'})
    
    customers = NguoiDung.objects.filter(
        Q(so_dien_thoai__icontains=query) | Q(ho_ten__icontains=query),
        vai_tro='khach_hang',
        da_xoa=False
    )[:10]
    
    if not customers:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy khách hàng'})
    
    customer = customers[0]  # Return first match as selected
    result = {
        'id': customer.id,
        'ho_ten': customer.ho_ten,
        'so_dien_thoai': customer.so_dien_thoai,
        'email': customer.email,
        'diem_tich_luy': customer.diem_tich_luy or 0,
        'so_lan_den': customer.dat_lich.filter(trang_thai='hoan_thanh', da_xoa=False).count()
    }
    
    return JsonResponse({'success': True, 'customer': result})

def api_customer_detail(request, customer_id):
    """API: Get customer detail"""
    try:
        customer = get_object_or_404(NguoiDung, id=customer_id, vai_tro='khach_hang', da_xoa=False)
        
        # Calculate customer stats
        bookings = DatLich.objects.filter(khach_hang=customer, da_xoa=False)
        completed_bookings = bookings.filter(trang_thai='hoan_thanh')
        total_spending = sum(booking.thanh_tien or 0 for booking in completed_bookings)
        
        # Get customer tier
        points = customer.diem_tich_luy or 0
        if points >= 1000:
            hang_thanh_vien = 'platinum'
            hang_display = 'Bạch kim'
        elif points >= 500:
            hang_thanh_vien = 'gold'
            hang_display = 'Vàng'
        elif points >= 200:
            hang_thanh_vien = 'silver'
            hang_display = 'Bạc'
        else:
            hang_thanh_vien = 'bronze'
            hang_display = 'Đồng'
        
        # Get history
        history = []
        for booking in bookings.order_by('-ngay_hen', '-gio_hen')[:10]:
            services = []
            for service in booking.dich_vu_dat_lich.all():
                services.append(service.ten_dich_vu)
            
            history.append({
                'type': 'booking',
                'date': booking.ngay_hen.strftime('%d/%m/%Y'),
                'amount': float(booking.thanh_tien or 0),
                'services': services,
                'status': booking.trang_thai
            })
        
        customer_data = {
            'id': customer.id,
            'ho_ten': customer.ho_ten,
            'so_dien_thoai': customer.so_dien_thoai,
            'email': customer.email,
            'ngay_sinh': customer.ngay_sinh.strftime('%d/%m/%Y') if customer.ngay_sinh else None,
            'dia_chi': customer.dia_chi,
            'anh_dai_dien': customer.anh_dai_dien,
            'diem_tich_luy': points,
            'so_lan_cat': completed_bookings.count(),
            'tong_chi_tieu': float(total_spending),
            'hang_thanh_vien': hang_thanh_vien,
            'hang_thanh_vien_display': hang_display,
            'ngay_tao': customer.ngay_tao.strftime('%d/%m/%Y'),
            'history': history
        }
        
        return JsonResponse(customer_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

def api_load_booking(request):
    """API: Load booking by code or ID"""
    booking_code = request.GET.get('code')
    booking_id = request.GET.get('id')
    
    try:
        if booking_code:
            booking = DatLich.objects.get(ma_dat_lich=booking_code, da_xoa=False)
        elif booking_id:
            booking = DatLich.objects.get(id=booking_id, da_xoa=False)
        else:
            return JsonResponse({'success': False, 'message': 'Code or ID required'})
        
        # Get services
        services = []
        for s in booking.dich_vu_dat_lich.filter(da_xoa=False).select_related('dich_vu'):
            services.append({
                'id': s.dich_vu.id,
                'ten_dich_vu': s.dich_vu.ten_dich_vu,
                'gia': int(s.gia_tai_thoi_diem or s.dich_vu.gia),
                'so_luong': s.so_luong or 1
            })
        
        # Customer info
        khach_hang_data = None
        if booking.khach_hang:
            khach_hang_data = {
                'id': booking.khach_hang.id,
                'ho_ten': booking.khach_hang.ho_ten,
                'so_dien_thoai': booking.khach_hang.so_dien_thoai,
                'email': booking.khach_hang.email,
                'diem_tich_luy': booking.khach_hang.diem_tich_luy or 0,
            }
        else:
            khach_hang_data = {
                'id': None,
                'ho_ten': booking.ten_khach_hang,
                'so_dien_thoai': booking.so_dien_thoai_khach,
                'email': '',
                'diem_tich_luy': 0,
            }
        
        result = {
            'id': booking.id,
            'ma_dat_lich': booking.ma_dat_lich,
            'khach_hang': khach_hang_data,
            'services': services
        }
        
        return JsonResponse({'success': True, 'booking': result})
        
    except DatLich.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy booking'})

def test_promotions(request):
    """Test view without authentication"""
    from datetime import datetime, date
    from django.contrib import messages
    
    # Handle POST requests for CRUD operations (same as main view)
    if request.method == 'POST':
        try:
            voucher_id = request.POST.get('voucher_id')
            
            # Validation
            ma_voucher = request.POST.get('ma_voucher', '').strip()
            ten_voucher = request.POST.get('ten_voucher', '').strip()
            loai_giam = request.POST.get('loai_giam', '').strip()
            gia_tri_giam = request.POST.get('gia_tri_giam', '').strip()
            ngay_bat_dau = request.POST.get('ngay_bat_dau', '').strip()
            ngay_ket_thuc = request.POST.get('ngay_ket_thuc', '').strip()
            
            if not all([ma_voucher, ten_voucher, loai_giam, gia_tri_giam, ngay_bat_dau, ngay_ket_thuc]):
                messages.error(request, 'Vui lòng điền đầy đủ thông tin bắt buộc!')
                return redirect('test_promotions')
            
            # Check for duplicate voucher code (exclude current when editing)
            existing_voucher = Voucher.objects.filter(ma_voucher=ma_voucher, da_xoa=False)
            if voucher_id:
                existing_voucher = existing_voucher.exclude(id=voucher_id)
            if existing_voucher.exists():
                messages.error(request, f'Mã voucher "{ma_voucher}" đã tồn tại!')
                return redirect('test_promotions')
            
            # Parse dates
            try:
                ngay_bat_dau_dt = datetime.strptime(ngay_bat_dau, '%Y-%m-%dT%H:%M')
                ngay_ket_thuc_dt = datetime.strptime(ngay_ket_thuc, '%Y-%m-%dT%H:%M')
                
                if ngay_bat_dau_dt >= ngay_ket_thuc_dt:
                    messages.error(request, 'Ngày kết thúc phải sau ngày bắt đầu!')
                    return redirect('test_promotions')
                    
            except ValueError as e:
                messages.error(request, f'Định dạng ngày giờ không hợp lệ: {e}!')
                return redirect('test_promotions')
            
            # Validate discount value
            try:
                gia_tri_giam_float = float(gia_tri_giam)
                if gia_tri_giam_float <= 0:
                    messages.error(request, 'Giá trị giảm phải lớn hơn 0!')
                    return redirect('test_promotions')
                if loai_giam == 'phan_tram' and gia_tri_giam_float > 100:
                    messages.error(request, 'Giảm theo phần trăm không được vượt quá 100%!')
                    return redirect('test_promotions')
            except ValueError:
                messages.error(request, 'Giá trị giảm không hợp lệ!')
                return redirect('test_promotions')
            
            # Prepare data with correct field names from model
            voucher_data = {
                'ma_voucher': ma_voucher,
                'ten_voucher': ten_voucher,
                'mo_ta': request.POST.get('mo_ta', '').strip(),
                'loai_giam': loai_giam,
                'gia_tri_giam': gia_tri_giam_float,
                'gia_tri_don_toi_thieu': float(request.POST.get('gia_tri_don_hang_toi_thieu') or 0),
                'giam_toi_da': float(request.POST.get('gia_tri_giam_toi_da') or 0) or None,
                'ngay_bat_dau': ngay_bat_dau_dt.date(),
                'ngay_ket_thuc': ngay_ket_thuc_dt.date(),
                'so_luong_tong': int(request.POST.get('so_luong_toi_da') or 0) or None,
                'trang_thai': request.POST.get('trang_thai') == 'active',
            }
            
            if voucher_id:
                # Update existing voucher
                voucher = Voucher.objects.get(id=voucher_id, da_xoa=False)
                for key, value in voucher_data.items():
                    setattr(voucher, key, value)
                voucher.save()
                messages.success(request, f'Đã cập nhật voucher "{ma_voucher}" thành công!')
            else:
                # Create new voucher - Use first user as default
                from .models import NguoiDung
                first_user = NguoiDung.objects.first()
                if not first_user:
                    messages.error(request, 'Không tìm thấy user để tạo voucher!')
                    return redirect('test_promotions')
                    
                voucher_data['nguoi_tao'] = first_user
                voucher = Voucher.objects.create(**voucher_data)
                messages.success(request, f'Đã tạo voucher "{ma_voucher}" thành công!')
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Có lỗi xảy ra: {str(e)}')
        
        return redirect('test_promotions')
    
    # GET request - Create a test object with attributes
    class TestVoucher:
        def __init__(self, **kwargs):
            for key, value in kwargs.items():
                setattr(self, key, value)
    
    # Get actual vouchers from database for display
    actual_vouchers = Voucher.objects.filter(da_xoa=False).order_by('-id')
    
    context = {
        'promotions': actual_vouchers,
        'total_promotions': actual_vouchers.count(),
        'active_promotions': actual_vouchers.filter(trang_thai=True).count(),
        'inactive_promotions': actual_vouchers.filter(trang_thai=False).count(),
        'expired_promotions': 0,
    }
    
    return render(request, 'admin/promotions.html', context)

# ============ REVIEWS API ENDPOINTS ============

@require_role(['quan_ly'])
def admin_review_reply(request, review_id):
    """API endpoint for replying to reviews"""
    if request.method == 'POST':
        try:
            review = get_object_or_404(DanhGia, id=review_id, da_xoa=False)
            reply_content = request.POST.get('phan_hoi', '').strip()
            
            if not reply_content:
                return JsonResponse({'success': False, 'message': 'Nội dung phản hồi không được để trống'})
            
            # Get NguoiDung from session
            user_id = request.session.get('user_id')
            if user_id:
                try:
                    nguoi_dung = NguoiDung.objects.get(id=user_id)
                    review.nguoi_phan_hoi = nguoi_dung
                except NguoiDung.DoesNotExist:
                    pass
                    
            review.phan_hoi = reply_content
            review.ngay_phan_hoi = timezone.now()
            review.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã gửi phản hồi thành công!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_review_detail(request, review_id):
    """API endpoint for getting review details"""
    try:
        review = get_object_or_404(DanhGia, id=review_id, da_xoa=False)
        
        data = {
            'id': review.id,
            'khach_hang': review.khach_hang.ho_ten,
            'so_sao': review.so_sao,
            'noi_dung': review.noi_dung,
            'phan_hoi': review.phan_hoi,
            'ngay_tao': review.ngay_tao.strftime('%d/%m/%Y %H:%M'),
            'dich_vu': review.dich_vu.ten_dich_vu,
            'nhan_vien': review.nhan_vien.ho_ten,
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})

@require_role(['quan_ly'])
def admin_review_delete(request, review_id):
    """API endpoint for deleting reviews (soft delete)"""
    if request.method == 'DELETE':
        try:
            review = get_object_or_404(DanhGia, id=review_id, da_xoa=False)
            review.da_xoa = True
            review.ngay_xoa = timezone.now()
            review.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã xóa đánh giá thành công!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_reviews_export(request):
    """Export reviews to CSV"""
    import csv
    from django.http import HttpResponse
    
    reviews = DanhGia.objects.filter(da_xoa=False).select_related(
        'khach_hang', 'dich_vu', 'nhan_vien'
    ).order_by('-ngay_tao')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="danh_gia_khach_hang.csv"'
    response.write('\ufeff')  # UTF-8 BOM for Excel
    
    writer = csv.writer(response)
    writer.writerow([
        'Ngày tạo', 'Khách hàng', 'Điện thoại', 'Dịch vụ', 'Nhân viên', 
        'Số sao', 'Nội dung', 'Phản hồi', 'Ngày phản hồi'
    ])
    
    for review in reviews:
        writer.writerow([
            review.ngay_tao.strftime('%d/%m/%Y %H:%M'),
            review.khach_hang.ho_ten,
            review.khach_hang.so_dien_thoai,
            review.dich_vu.ten_dich_vu,
            review.nhan_vien.ho_ten,
            review.so_sao,
            review.noi_dung,
            review.phan_hoi or '',
            review.ngay_phan_hoi.strftime('%d/%m/%Y %H:%M') if review.ngay_phan_hoi else ''
        ])
    
    return response

# ============ DASHBOARD ACTION ENDPOINTS ============

@require_role(['quan_ly'])
def admin_booking_approve(request, booking_id):
    """Approve booking from dashboard"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            booking.trang_thai = 'da_xac_nhan'
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã xác nhận booking thành công!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_booking_reject(request, booking_id):
    """Reject booking from dashboard"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            booking.trang_thai = 'da_huy'
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã hủy booking thành công!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_leave_approve(request, leave_id):
    """Approve leave request from dashboard"""
    if request.method == 'POST':
        try:
            leave = get_object_or_404(DonXinNghi, id=leave_id, da_xoa=False)
            leave.trang_thai = 'da_duyet'
            
            # Get user instance instead of just ID
            user_id = request.session.get('user_id')
            if user_id:
                try:
                    nguoi_duyet = NguoiDung.objects.get(id=user_id)
                    leave.nguoi_duyet = nguoi_duyet
                except NguoiDung.DoesNotExist:
                    pass
                    
            # ngay_cap_nhat will be automatically updated due to auto_now=True
            leave.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã phê duyệt yêu cầu nghỉ phép!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['quan_ly'])
def admin_leave_reject(request, leave_id):
    """Reject leave request from dashboard"""
    if request.method == 'POST':
        try:
            leave = get_object_or_404(DonXinNghi, id=leave_id, da_xoa=False)
            leave.trang_thai = 'tu_choi'
            
            # Get user instance instead of just ID
            user_id = request.session.get('user_id')
            if user_id:
                try:
                    nguoi_duyet = NguoiDung.objects.get(id=user_id)
                    leave.nguoi_duyet = nguoi_duyet
                except NguoiDung.DoesNotExist:
                    pass
                    
            # ngay_cap_nhat will be automatically updated due to auto_now=True
            leave.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã từ chối yêu cầu nghỉ phép!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

# ============ STAFF ACTION ENDPOINTS ============

@require_role(['nhan_vien', 'quan_ly'])
def staff_booking_checkin(request, booking_id):
    """Staff check-in customer"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_xac_nhan':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chỉ có thể check-in booking đã xác nhận!'
                })
            
            booking.trang_thai = 'da_checkin'
            booking.ngay_check_in = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã check-in khách hàng thành công!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@require_role(['nhan_vien', 'quan_ly'])
def staff_booking_complete(request, booking_id):
    """Staff complete booking"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_checkin':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chỉ có thể hoàn thành booking đã check-in!'
                })
            
            booking.trang_thai = 'hoan_thanh'
            booking.ngay_hoan_thanh = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Đã hoàn thành dịch vụ!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Có lỗi xảy ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

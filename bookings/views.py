"""
Views for bookings app (Bookings, Invoices, POS)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Sum, Q, F, Max
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta, time
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Import decorators from core
from core.decorators import require_auth, require_role

# Import models
from barbershop.models import (
    NguoiDung,
    DatLich,
    DichVu,
    DichVuDatLich,
    HoaDon,
    ChiTietHoaDon,
    Voucher,
    ThongTinNhanVien
)

# ============ VIEWS - Bạn sẽ copy code vào đây ============
# Các views sẽ được di chuyển vào đây:
# - admin_bookings
@require_role(['quan_ly'])
def admin_bookings(request):
    """Booking Management"""
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    staff_id = request.GET.get('staff_id')
    search = request.GET.get('search')
    
    # Debug logging
    print(f"DEBUG - Filter parameters: from_date={from_date}, to_date={to_date}, staff_id={staff_id}, search={search}")
    
    # Base queryset
    bookings = DatLich.objects.filter(
        da_xoa=False
    ).select_related('khach_hang', 'nhan_vien').prefetch_related(
        'dich_vu_dat_lich__dich_vu'
    ).order_by('-ngay_hen', '-gio_hen')
    
    # Apply filters
    original_count = bookings.count()
    
    if from_date:
        bookings = bookings.filter(ngay_hen__gte=from_date)
        print(f"DEBUG - After from_date filter ({from_date}): {bookings.count()} bookings")
    
    if to_date:
        bookings = bookings.filter(ngay_hen__lte=to_date)
        print(f"DEBUG - After to_date filter ({to_date}): {bookings.count()} bookings")
    
    if staff_id and staff_id.strip():
        bookings = bookings.filter(nhan_vien_id=staff_id.strip())
        print(f"DEBUG - After staff filter ({staff_id.strip()}): {bookings.count()} bookings")
    
    if search and search.strip() and search.strip().lower() != 'none':
        from django.db.models import Q
        search_term = search.strip()
        bookings = bookings.filter(
            Q(so_dien_thoai_khach__icontains=search_term) |
            Q(ten_khach_hang__icontains=search_term) |
            Q(ma_dat_lich__icontains=search_term)
        )
        print(f"DEBUG - After search filter ({search_term}): {bookings.count()} bookings")
    
    print(f"DEBUG - Total bookings after all filters: {bookings.count()} (original: {original_count})")
    
    # Get all bookings for statistics (without pagination)
    all_bookings = bookings
    
    # Statistics
    total_bookings = all_bookings.count()
    pending_count = all_bookings.filter(trang_thai='cho_xac_nhan').count()
    confirmed_count = all_bookings.filter(trang_thai='da_xac_nhan').count()
    in_progress_count = all_bookings.filter(trang_thai='da_checkin').count()
    completed_count = all_bookings.filter(trang_thai='hoan_thanh').count()
    cancelled_count = all_bookings.filter(trang_thai='da_huy').count()
    
    # Get staff list for filter
    staff_list = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Get customers for modal
    customers = NguoiDung.objects.filter(
        vai_tro='khach_hang',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Get services for modal
    services = DichVu.objects.filter(
        trang_thai=True,
        da_xoa=False
    ).select_related('danh_muc').order_by('thu_tu', 'ten_dich_vu')
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(bookings, 25)  # 25 bookings per page
    page_number = request.GET.get('page')
    bookings_page = paginator.get_page(page_number)
    
    context = {
        'bookings': bookings_page,
        'total_bookings': total_bookings,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'staff_list': staff_list,
        'customers': customers,
        'services': services,
        'from_date': from_date,
        'to_date': to_date,
        'staff_id': staff_id,
        'search': search,
    }
    return render(request, 'admin/bookings.html', context)
# - admin_bookings_create
@require_role(['quan_ly'])
def admin_bookings_create(request):
    """Create Booking with proper validation"""
    if request.method == 'POST':
        try:
            # Validate required fields
            customer_id = request.POST.get('customer_id')
            staff_id = request.POST.get('staff_id') 
            booking_date = request.POST.get('booking_date')
            booking_time = request.POST.get('booking_time')
            services = request.POST.getlist('services[]')  # Note: services[], not service_ids
            
            if not all([customer_id, staff_id, booking_date, booking_time]):
                from django.contrib import messages
                messages.error(request, 'Vui lÃ²ng Äiá»n Äáº§y Äá»§ thÃ´ng tin báº¯t buá»c!')
                return redirect(request.get_full_path())
            
            if not services:
                from django.contrib import messages  
                messages.error(request, 'Vui lÃ²ng chá»n Ã­t nháº¥t má»t dá»ch vá»¥!')
                return redirect(request.get_full_path())
            
            # Calculate total amount and generate booking code
            total_amount = 0
            for service_id in services:
                service = DichVu.objects.get(id=service_id)
                total_amount += service.gia
            
            # Generate unique booking code
            import random, string
            booking_code = 'BK' + ''.join(random.choices(string.digits, k=6))
            while DatLich.objects.filter(ma_dat_lich=booking_code).exists():
                booking_code = 'BK' + ''.join(random.choices(string.digits, k=6))
            
            # Get customer info
            customer = NguoiDung.objects.get(id=customer_id)
            
            # Create booking
            booking = DatLich.objects.create(
                ma_dat_lich=booking_code,
                khach_hang=customer,
                ten_khach_hang=customer.ho_ten,
                so_dien_thoai_khach=customer.so_dien_thoai,
                email_khach=customer.email or None,
                nhan_vien_id=staff_id if staff_id != 'auto' else None,
                ngay_hen=booking_date,
                gio_hen=booking_time,
                tong_tien=total_amount,
                thanh_tien=total_amount,
                trang_thai='da_xac_nhan',
                ghi_chu=request.POST.get('notes', ''),
                da_xoa=False
            )
            
            # Add services
            for service_id in services:
                service = DichVu.objects.get(id=service_id)
                DichVuDatLich.objects.create(
                    dat_lich=booking,
                    dich_vu=service,
                    ten_dich_vu=service.ten_dich_vu,
                    gia=service.gia,
                    so_luong=1,
                    thanh_tien=service.gia
                )
            
            from django.contrib import messages
            messages.success(request, f'ÄÃ£ táº¡o lá»ch háº¹n {booking_code} thÃ nh cÃ´ng!')
            return redirect('bookings:admin_bookings')
            
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Lá»i táº¡o lá»ch háº¹n: {str(e)}')
            return redirect(request.get_full_path())
    
    # Get data for form
    customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False).order_by('ho_ten')
    staff_list = NguoiDung.objects.filter(vai_tro='nhan_vien', da_xoa=False, trang_thai=True).order_by('ho_ten')
    services = DichVu.objects.filter(da_xoa=False, trang_thai=True).order_by('ten_dich_vu')
    
    # Generate time slots (8:00 - 20:00)
    from datetime import time
    time_slots = []
    for hour in range(8, 21):  # 8AM to 8PM
        for minute in [0, 30]:  # Every 30 minutes
            if hour == 20 and minute == 30:  # Don't go past 8PM
                break
            slot_time = time(hour, minute)
            time_slots.append(slot_time.strftime('%H:%M'))
    
    # Get selected customer ID from URL parameter
    selected_customer_id = request.GET.get('customer_id')
    
    from datetime import date
    context = {
        'customers': customers,
        'staff': staff_list,  # Changed from 'staff' to 'staff_list'
        'staff_list': staff_list,  # Add both for compatibility
        'services': services,
        'time_slots': time_slots,
        'today': date.today().isoformat(),
        'selected_customer_id': selected_customer_id,
    }
    return render(request, 'admin/bookings-create.html', context)
# - admin_booking_detail
@require_role(['quan_ly'])
def admin_booking_detail(request, booking_id):
    """Booking Detail and Management"""
    booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            new_status = request.POST.get('trang_thai')
            booking.trang_thai = new_status
            
            # Update timestamps based on status
            if new_status == 'da_checkin':
                booking.ngay_check_in = timezone.now()
            elif new_status == 'hoan_thanh':
                booking.ngay_hoan_thanh = timezone.now()
            
            booking.save()
            
        elif action == 'cancel':
            booking.trang_thai = 'da_huy'
            booking.ly_do_huy = request.POST.get('ly_do_huy')
            booking.save()
            
            # TODO: Send notification to customer
            
        return redirect('admin_booking_detail', booking_id=booking_id)
    
    # GET request - show booking details
    services = booking.dich_vu_dat_lich.all().select_related('dich_vu')
    
    context = {
        'booking': booking,
        'services': services,
    }
    return render(request, 'admin/booking-detail.html', context)
# - admin_booking_cancel
@require_role(['quan_ly'])
def admin_booking_cancel(request, booking_id):
    """API to cancel booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai in ['hoan_thanh', 'da_huy']:
                return JsonResponse({
                    'success': False, 
                    'message': 'KhÃ´ng thá» há»§y lá»ch háº¹n ÄÃ£ hoÃ n thÃ nh hoáº·c ÄÃ£ há»§y'
                })
            
            booking.trang_thai = 'da_huy'
            booking.ngay_huy = timezone.now()
            booking.ly_do_huy = request.POST.get('reason', 'Há»§y tá»« admin')
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'ÄÃ£ há»§y lá»ch háº¹n thÃ nh cÃ´ng'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lá»i: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - admin_booking_checkin
@require_role(['quan_ly'])
def admin_booking_checkin(request, booking_id):
    """API to check-in booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai != 'da_xac_nhan':
                return JsonResponse({
                    'success': False,
                    'message': 'Chá» cÃ³ thá» check-in lá»ch háº¹n ÄÃ£ xÃ¡c nháº­n'
                })
            
            booking.trang_thai = 'da_checkin'
            booking.ngay_check_in = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'ÄÃ£ check-in thÃ nh cÃ´ng'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lá»i: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - admin_booking_complete
@require_role(['quan_ly'])
def admin_booking_complete(request, booking_id):
    """API to complete booking"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            
            if booking.trang_thai != 'da_checkin':
                return JsonResponse({
                    'success': False,
                    'message': 'Chá» cÃ³ thá» hoÃ n thÃ nh lá»ch háº¹n ÄÃ£ check-in'
                })
            
            booking.trang_thai = 'hoan_thanh'
            booking.ngay_hoan_thanh = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True,
                'message': 'ÄÃ£ hoÃ n thÃ nh lá»ch háº¹n'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lá»i: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

# - admin_bookings_export
@require_role(['quan_ly'])
def admin_bookings_export(request):
    """Export bookings to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset
        bookings = DatLich.objects.filter(
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien').prefetch_related(
            'dich_vu_dat_lich__dich_vu'
        ).order_by('-ngay_hen', '-gio_hen')
        
        # Apply filters
        if from_date:
            bookings = bookings.filter(ngay_hen__gte=from_date)
        
        if to_date:
            bookings = bookings.filter(ngay_hen__lte=to_date)
        
        if staff_id:
            bookings = bookings.filter(nhan_vien_id=staff_id)
        
        if search:
            from django.db.models import Q
            bookings = bookings.filter(
                Q(so_dien_thoai_khach__icontains=search) |
                Q(ten_khach_hang__icontains=search) |
                Q(ma_dat_lich__icontains=search)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sÃ¡ch Äáº·t lá»ch"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'MÃ£ Äáº·t lá»ch', 'KhÃ¡ch hÃ ng', 'Sá» Äiá»n thoáº¡i', 'NhÃ¢n viÃªn',
            'NgÃ y háº¹n', 'Giá» háº¹n', 'Dá»ch vá»¥', 'Tá»ng tiá»n', 'Tráº¡ng thÃ¡i', 
            'NgÃ y táº¡o', 'Ghi chÃº'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        status_map = {
            'cho_xac_nhan': 'Chá» xÃ¡c nháº­n',
            'da_xac_nhan': 'ÄÃ£ xÃ¡c nháº­n',
            'da_checkin': 'Äang phá»¥c vá»¥',
            'hoan_thanh': 'HoÃ n thÃ nh',
            'da_huy': 'ÄÃ£ há»§y',
            'khong_den': 'KhÃ´ng Äáº¿n'
        }
        
        for row, booking in enumerate(bookings, 2):
            services = ', '.join([s.ten_dich_vu for s in booking.dich_vu_dat_lich.all()])
            
            ws.cell(row=row, column=1, value=booking.ma_dat_lich)
            ws.cell(row=row, column=2, value=booking.ten_khach_hang)
            ws.cell(row=row, column=3, value=booking.so_dien_thoai_khach)
            ws.cell(row=row, column=4, value=booking.nhan_vien.ho_ten if booking.nhan_vien else '')
            ws.cell(row=row, column=5, value=booking.ngay_hen.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=booking.gio_hen.strftime('%H:%M'))
            ws.cell(row=row, column=7, value=services)
            ws.cell(row=row, column=8, value=float(booking.thanh_tien))
            ws.cell(row=row, column=9, value=status_map.get(booking.trang_thai, booking.trang_thai))
            ws.cell(row=row, column=10, value=booking.ngay_tao.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=11, value=booking.ghi_chu or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'danh_sach_dat_lich_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'Cáº§n cÃ i Äáº·t openpyxl Äá» xuáº¥t Excel: pip install openpyxl'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Lá»i xuáº¥t Excel: {str(e)}'
        })
# - admin_booking_approve (dashboard action)
@require_role(['quan_ly'])
def admin_booking_approve(request, booking_id):
    """Approve booking from dashboard"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            booking.trang_thai = 'da_xac_nhan'
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ xÃ¡c nháº­n booking thÃ nh cÃ´ng!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - admin_booking_reject (dashboard action)
@require_role(['quan_ly'])
def admin_booking_reject(request, booking_id):
    """Reject booking from dashboard"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(DatLich, id=booking_id, da_xoa=False)
            booking.trang_thai = 'da_huy'
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ há»§y booking thÃ nh cÃ´ng!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - admin_invoices
@require_role(['quan_ly'])
def admin_invoices(request):
    """Invoice Management with filters and stats"""
    from datetime import date
    from django.db.models import Sum, Avg, Count
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    payment_method = request.GET.get('payment_method')
    staff_id = request.GET.get('staff_id')
    customer_type = request.GET.get('customer_type')
    search = request.GET.get('search')
    sort_by = request.GET.get('sort_by', '-ngay_thanh_toan')
    
    # Debug logging
    print(f"DEBUG Invoice - Filter parameters: from_date={from_date}, to_date={to_date}, payment_method={payment_method}, staff_id={staff_id}, search={search}")
    
    # Base queryset
    invoices = HoaDon.objects.filter(
        da_xoa=False
    ).select_related('khach_hang', 'nhan_vien', 'nguoi_tao', 'dat_lich').prefetch_related(
        'chi_tiet__dich_vu'
    )
    
    # Apply filters
    original_count = invoices.count()
    
    if from_date and from_date.strip() and from_date.strip().lower() != 'none':
        invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        print(f"DEBUG Invoice - After from_date filter ({from_date}): {invoices.count()} invoices")
    
    if to_date and to_date.strip() and to_date.strip().lower() != 'none':
        invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        print(f"DEBUG Invoice - After to_date filter ({to_date}): {invoices.count()} invoices")
    
    if payment_method and payment_method.strip():
        invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        print(f"DEBUG Invoice - After payment_method filter ({payment_method}): {invoices.count()} invoices")
    
    if staff_id and staff_id.strip():
        invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        print(f"DEBUG Invoice - After staff filter ({staff_id}): {invoices.count()} invoices")
    
    if search and search.strip() and search.strip().lower() != 'none':
        from django.db.models import Q
        search_term = search.strip()
        invoices = invoices.filter(
            Q(ma_hoa_don__icontains=search_term) |
            Q(ten_khach_hang__icontains=search_term) |
            Q(so_dien_thoai_khach__icontains=search_term)
        )
        print(f"DEBUG Invoice - After search filter ({search_term}): {invoices.count()} invoices")
    
    # Apply sorting
    valid_sort_fields = [
        'ngay_thanh_toan', '-ngay_thanh_toan',
        'thanh_tien', '-thanh_tien', 
        'ma_hoa_don', '-ma_hoa_don',
        'ten_khach_hang', '-ten_khach_hang'
    ]
    
    if sort_by and sort_by in valid_sort_fields:
        invoices = invoices.order_by(sort_by)
    else:
        invoices = invoices.order_by('-ngay_thanh_toan')
    
    print(f"DEBUG Invoice - Total invoices after all filters: {invoices.count()} (original: {original_count})")
    
    # Get all invoices for statistics (without pagination)
    all_invoices = invoices
    
    # Statistics
    today = date.today()
    
    total_invoices = all_invoices.count()
    total_revenue = all_invoices.aggregate(total=Sum('thanh_tien'))['total'] or 0
    today_invoices = all_invoices.filter(ngay_thanh_toan__date=today).count()
    today_revenue = all_invoices.filter(ngay_thanh_toan__date=today).aggregate(total=Sum('thanh_tien'))['total'] or 0
    average_invoice = all_invoices.aggregate(avg=Avg('thanh_tien'))['avg'] or 0
    
    # Payment method stats
    payment_methods = all_invoices.values('phuong_thuc_thanh_toan').annotate(
        count=Count('id'),
        total=Sum('thanh_tien')
    ).order_by('-total')
    
    # Get staff list for filter
    staff_list = NguoiDung.objects.filter(
        vai_tro='nhan_vien',
        trang_thai=True,
        da_xoa=False
    ).order_by('ho_ten')
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(invoices, 25)  # 25 invoices per page
    page_number = request.GET.get('page')
    invoices_page = paginator.get_page(page_number)
    
    # Calculate page total
    page_total = sum(invoice.thanh_tien for invoice in invoices_page.object_list)
    
    context = {
        'invoices': invoices_page,
        'total_invoices': total_invoices,
        'total_revenue': total_revenue,
        'today_invoices': today_invoices,
        'today_revenue': today_revenue,
        'average_invoice': average_invoice,
        'payment_methods': payment_methods,
        'staff_list': staff_list,
        'page_total': page_total,
        'from_date': from_date,
        'to_date': to_date,
        'payment_method': payment_method,
        'staff_id': staff_id,
        'customer_type': customer_type,
        'search': search,
        'sort_by': sort_by,
    }
    return render(request, 'admin/invoices.html', context)

# - admin_invoices_export_excel
@require_role(['quan_ly'])
def admin_invoices_export_excel(request):
    """Export invoices to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from datetime import datetime
        
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        payment_method = request.GET.get('payment_method')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset
        invoices = HoaDon.objects.filter(
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien', 'nguoi_tao').prefetch_related(
            'chi_tiet__dich_vu'
        ).order_by('-ngay_thanh_toan')
        
        # Apply filters (same logic as main view)
        if from_date and from_date.strip() and from_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        
        if to_date and to_date.strip() and to_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        
        if payment_method and payment_method.strip():
            invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        
        if staff_id and staff_id.strip():
            invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        
        if search and search.strip() and search.strip().lower() != 'none':
            from django.db.models import Q
            search_term = search.strip()
            invoices = invoices.filter(
                Q(ma_hoa_don__icontains=search_term) |
                Q(ten_khach_hang__icontains=search_term) |
                Q(so_dien_thoai_khach__icontains=search_term)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sÃ¡ch hÃ³a ÄÆ¡n"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'MÃ£ hÃ³a ÄÆ¡n', 'KhÃ¡ch hÃ ng', 'Sá» Äiá»n thoáº¡i', 'NhÃ¢n viÃªn',
            'NgÃ y thanh toÃ¡n', 'Dá»ch vá»¥', 'Táº¡m tÃ­nh', 'Giáº£m giÃ¡', 'ThÃ nh tiá»n', 
            'PhÆ°Æ¡ng thá»©c TT', 'Ghi chÃº'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Payment method mapping
        payment_method_map = {
            'tien_mat': 'Tiá»n máº·t',
            'chuyen_khoan': 'Chuyá»n khoáº£n',
            'vi_dien_tu': 'VÃ­ Äiá»n tá»­',
            'the': 'Tháº»'
        }
        
        # Data rows
        for row, invoice in enumerate(invoices, 2):
            services = ', '.join([ct.ten_dich_vu for ct in invoice.chi_tiet.all()])
            
            ws.cell(row=row, column=1, value=invoice.ma_hoa_don)
            ws.cell(row=row, column=2, value=invoice.ten_khach_hang)
            ws.cell(row=row, column=3, value=invoice.so_dien_thoai_khach)
            ws.cell(row=row, column=4, value=invoice.nhan_vien.ho_ten if invoice.nhan_vien else '')
            ws.cell(row=row, column=5, value=invoice.ngay_thanh_toan.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=6, value=services)
            ws.cell(row=row, column=7, value=float(invoice.tam_tinh))
            ws.cell(row=row, column=8, value=float(invoice.tien_giam_gia))
            ws.cell(row=row, column=9, value=float(invoice.thanh_tien))
            ws.cell(row=row, column=10, value=payment_method_map.get(invoice.phuong_thuc_thanh_toan, invoice.phuong_thuc_thanh_toan))
            ws.cell(row=row, column=11, value=invoice.ghi_chu or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'danh_sach_hoa_don_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lá»i xuáº¥t Excel: {str(e)}'
        })

# - admin_invoices_export_pdf
@require_role(['quan_ly'])
def admin_invoices_export_pdf(request):
    """Export invoices to PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        import io
        
        # Get filter parameters (same as Excel export)
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        payment_method = request.GET.get('payment_method')
        staff_id = request.GET.get('staff_id')
        search = request.GET.get('search')
        
        # Base queryset with filters (same logic)
        invoices = HoaDon.objects.filter(da_xoa=False).select_related('khach_hang', 'nhan_vien').prefetch_related('chi_tiet__dich_vu').order_by('-ngay_thanh_toan')
        
        # Apply filters
        if from_date and from_date.strip() and from_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__gte=from_date.strip())
        if to_date and to_date.strip() and to_date.strip().lower() != 'none':
            invoices = invoices.filter(ngay_thanh_toan__date__lte=to_date.strip())
        if payment_method and payment_method.strip():
            invoices = invoices.filter(phuong_thuc_thanh_toan=payment_method.strip())
        if staff_id and staff_id.strip():
            invoices = invoices.filter(nhan_vien_id=staff_id.strip())
        if search and search.strip() and search.strip().lower() != 'none':
            from django.db.models import Q
            search_term = search.strip()
            invoices = invoices.filter(Q(ma_hoa_don__icontains=search_term) | Q(ten_khach_hang__icontains=search_term) | Q(so_dien_thoai_khach__icontains=search_term))
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Content
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=1)
        title = Paragraph("DANH SÃCH HÃA ÄÆ N", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['MÃ£ HÄ', 'KhÃ¡ch hÃ ng', 'SÄT', 'NhÃ¢n viÃªn', 'NgÃ y TT', 'ThÃ nh tiá»n', 'PT Thanh toÃ¡n']]
        
        payment_method_map = {'tien_mat': 'Tiá»n máº·t', 'chuyen_khoan': 'CK', 'vi_dien_tu': 'VÃ­', 'the': 'Tháº»'}
        
        for invoice in invoices[:50]:  # Limit for PDF
            data.append([
                invoice.ma_hoa_don,
                invoice.ten_khach_hang[:15] + '...' if len(invoice.ten_khach_hang) > 15 else invoice.ten_khach_hang,
                invoice.so_dien_thoai_khach,
                invoice.nhan_vien.ho_ten[:10] + '...' if invoice.nhan_vien and len(invoice.nhan_vien.ho_ten) > 10 else (invoice.nhan_vien.ho_ten if invoice.nhan_vien else ''),
                invoice.ngay_thanh_toan.strftime('%d/%m/%Y'),
                f"{invoice.thanh_tien:,.0f}Ä",
                payment_method_map.get(invoice.phuong_thuc_thanh_toan, invoice.phuong_thuc_thanh_toan)
            ])
        
        # Create table
        table = Table(data, colWidths=[1*inch, 1.5*inch, 1*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f'danh_sach_hoa_don_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': 'Cáº§n cÃ i Äáº·t reportlab Äá» xuáº¥t PDF: pip install reportlab'
        })
    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'message': f'Lá»i xuáº¥t PDF: {str(e)}'
        })

# - staff_pos
@require_role(['nhan_vien', 'quan_ly'])
def staff_pos(request):
    """POS - Point of Sale System"""
    user_id = request.session.get('user_id')
    
    if request.method == 'POST':
        # Process payment
        try:
            # Get customer info
            customer_type = request.POST.get('customer_type', 'walk-in')
            customer_id = request.POST.get('customer_id')
            customer_name = request.POST.get('customer_name', '')
            customer_phone = request.POST.get('customer_phone', '')
            booking_id = request.POST.get('booking_id')
            
            # Get services from cart (JSON string)
            import json
            cart_data = request.POST.get('cart_data', '[]')
            cart = json.loads(cart_data)
            
            if not cart:
                return JsonResponse({'success': False, 'message': 'Vui lÃ²ng chá»n Ã­t nháº¥t má»t dá»ch vá»¥!'})
            
            # Get payment info
            payment_method = request.POST.get('payment_method', 'tien_mat')
            voucher_id = request.POST.get('voucher_id')
            points_used = int(request.POST.get('points_used', 0) or 0)
            stylist_id = request.POST.get('stylist_id', user_id)
            
            # Calculate totals
            tam_tinh = sum(int(item['price']) * int(item['quantity']) for item in cart)
            tien_giam_gia = 0
            
            # Apply voucher
            if voucher_id:
                voucher = Voucher.objects.filter(id=voucher_id, trang_thai='hoat_dong', da_xoa=False).first()
                if voucher:
                    if voucher.loai_giam == 'tien':
                        tien_giam_gia += voucher.gia_tri_giam
                    else:  # phan_tram
                        tien_giam_gia += int(tam_tinh * voucher.gia_tri_giam / 100)
            
            # Apply points (1 point = 1,000 VND)
            if points_used > 0:
                tien_giam_gia += points_used * 1000
            
            thanh_tien = tam_tinh - tien_giam_gia
            
            # Create or link booking
            if booking_id:
                # From existing booking
                booking = DatLich.objects.get(id=booking_id)
                booking.trang_thai = 'hoan_thanh'
                booking.ngay_hoan_thanh = timezone.now()
                booking.save()
            else:
                # Create new booking
                booking = DatLich.objects.create(
                    khach_hang_id=customer_id if customer_id else None,
                    ten_khach_hang=customer_name,
                    so_dien_thoai_khach=customer_phone,
                    nhan_vien_id=stylist_id,
                    ngay_hen=timezone.now().date(),
                    gio_hen=timezone.now().time(),
                    loai_dat_lich='walk_in',
                    trang_thai='hoan_thanh',
                    ngay_check_in=timezone.now(),
                    ngay_hoan_thanh=timezone.now(),
                    tong_tien=tam_tinh,
                    tien_giam_gia=tien_giam_gia,
                    thanh_tien=thanh_tien,
                    da_xoa=False
                )
                
                # Add services to booking
                for item in cart:
                    DichVuDatLich.objects.create(
                        dat_lich=booking,
                        dich_vu_id=item['id'],
                        so_luong=item['quantity'],
                        gia_tai_thoi_diem=item['price'],
                        thanh_tien=int(item['price']) * int(item['quantity']),
                        da_xoa=False
                    )
            
            # Create invoice
            invoice = HoaDon.objects.create(
                dat_lich=booking,
                khach_hang_id=customer_id if customer_id else None,
                nhan_vien_id=stylist_id,
                phuong_thuc_thanh_toan=payment_method,
                voucher_id=voucher_id if voucher_id else None,
                diem_su_dung=points_used,
                tam_tinh=tam_tinh,
                tien_giam_gia=tien_giam_gia,
                thanh_tien=thanh_tien,
                trang_thai='da_thanh_toan',
                ngay_thanh_toan=timezone.now(),
                da_xoa=False
            )
            
            # Update customer points
            if customer_id:
                customer = NguoiDung.objects.get(id=customer_id)
                # Deduct used points
                if points_used > 0:
                    customer.diem_tich_luy = max(0, customer.diem_tich_luy - points_used)
                # Add new points (1% of thanh_tien)
                new_points = int(thanh_tien / 1000 * 0.01)
                customer.diem_tich_luy += new_points
                customer.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Thanh toÃ¡n thÃ nh cÃ´ng!',
                'invoice_id': invoice.id,
                'ma_hoa_don': invoice.ma_hoa_don
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    # GET request - show POS interface
    try:
        services = DichVu.objects.filter(da_xoa=False, trang_thai=True).select_related('danh_muc')
    except Exception as e:
        print(f"Error loading services: {e}")
        services = DichVu.objects.filter(da_xoa=False).select_related('danh_muc')
    
    try:
        customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False)[:50]  # Recent customers
    except Exception as e:
        print(f"Error loading customers: {e}")
        customers = []
    
    try:
        today_bookings = DatLich.objects.filter(
            ngay_hen=timezone.now().date(),
            trang_thai__in=['da_xac_nhan', 'da_checkin'],
            da_xoa=False
        ).select_related('khach_hang', 'nhan_vien')
    except Exception as e:
        print(f"Error loading today_bookings: {e}")
        today_bookings = []
    
    try:
        # Fixed: Use boolean True instead of string 'hoat_dong'
        vouchers = Voucher.objects.filter(
            trang_thai=True,
            ngay_bat_dau__lte=timezone.now(),
            ngay_ket_thuc__gte=timezone.now(),
            da_xoa=False
        )
    except Exception as e:
        print(f"Error loading vouchers: {e}")
        vouchers = []
    
    try:
        staff = NguoiDung.objects.filter(vai_tro='nhan_vien', da_xoa=False, trang_thai=True)
    except Exception as e:
        print(f"Error loading staff: {e}")
        staff = []
    
    context = {
        'services': services,
        'customers': customers,
        'today_bookings': today_bookings,
        'vouchers': vouchers,
        'staff': staff,
        'current_user_id': user_id,
    }
    return render(request, 'staff/pos.html', context)
# - staff_bookings_create
@require_role(['nhan_vien', 'quan_ly'])
def staff_bookings_create(request):
    """Create Booking (Staff)"""
    if request.method == 'POST':
        user_id = request.session.get('user_id')
        
        booking = DatLich.objects.create(
            khach_hang_id=request.POST.get('khach_hang_id'),
            nhan_vien_id=user_id,
            ngay_hen=request.POST.get('ngay_hen'),
            gio_hen=request.POST.get('gio_hen'),
            trang_thai='da_xac_nhan',
            ghi_chu=request.POST.get('ghi_chu', ''),
            da_xoa=False
        )
        
        service_ids = request.POST.getlist('service_ids')
        for service_id in service_ids:
            DichVuDatLich.objects.create(
                dat_lich=booking,
                dich_vu_id=service_id,
                da_xoa=False
            )
        
        return redirect('bookings:staff_today_bookings')
    
    customers = NguoiDung.objects.filter(vai_tro='khach_hang', da_xoa=False)
    services = DichVu.objects.filter(da_xoa=False, trang_thai=True)
    
    context = {
        'customers': customers,
        'services': services,
    }
    return render(request, 'staff/bookings-create.html', context)
# - api_search_customer
def api_search_customer(request):
    """API: Search customer by phone or name"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'success': False, 'message': 'Query required'})
    
    customers = NguoiDung.objects.filter(
        Q(so_dien_thoai__icontains=query) | Q(ho_ten__icontains=query),
        vai_tro='khach_hang',
        da_xoa=False
    )[:10]
    
    if not customers:
        return JsonResponse({'success': False, 'message': 'KhÃ´ng tÃ¬m tháº¥y khÃ¡ch hÃ ng'})
    
    customer = customers[0]  # Return first match as selected
    result = {
        'id': customer.id,
        'ho_ten': customer.ho_ten,
        'so_dien_thoai': customer.so_dien_thoai,
        'email': customer.email,
        'diem_tich_luy': customer.diem_tich_luy or 0,
        'so_lan_den': customer.dat_lich.filter(trang_thai='hoan_thanh', da_xoa=False).count()
    }
    
    return JsonResponse({'success': True, 'customer': result})
# - api_load_booking
def api_load_booking(request):
    """API: Load booking by code or ID"""
    booking_code = request.GET.get('code')
    booking_id = request.GET.get('id')
    
    try:
        if booking_code:
            booking = DatLich.objects.get(ma_dat_lich=booking_code, da_xoa=False)
        elif booking_id:
            booking = DatLich.objects.get(id=booking_id, da_xoa=False)
        else:
            return JsonResponse({'success': False, 'message': 'Code or ID required'})
        
        # Get services
        services = []
        for s in booking.dich_vu_dat_lich.filter(da_xoa=False).select_related('dich_vu'):
            services.append({
                'id': s.dich_vu.id,
                'ten_dich_vu': s.dich_vu.ten_dich_vu,
                'gia': int(s.gia_tai_thoi_diem or s.dich_vu.gia),
                'so_luong': s.so_luong or 1
            })
        
        # Customer info
        khach_hang_data = None
        if booking.khach_hang:
            khach_hang_data = {
                'id': booking.khach_hang.id,
                'ho_ten': booking.khach_hang.ho_ten,
                'so_dien_thoai': booking.khach_hang.so_dien_thoai,
                'email': booking.khach_hang.email,
                'diem_tich_luy': booking.khach_hang.diem_tich_luy or 0,
            }
        else:
            khach_hang_data = {
                'id': None,
                'ho_ten': booking.ten_khach_hang,
                'so_dien_thoai': booking.so_dien_thoai_khach,
                'email': '',
                'diem_tich_luy': 0,
            }
        
        result = {
            'id': booking.id,
            'ma_dat_lich': booking.ma_dat_lich,
            'khach_hang': khach_hang_data,
            'services': services
        }
        
        return JsonResponse({'success': True, 'booking': result})
        
    except DatLich.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'KhÃ´ng tÃ¬m tháº¥y booking'})

# - api_booking_confirm
@require_role(['nhan_vien', 'quan_ly'])
def api_booking_confirm(request, booking_id):
    """Confirm booking via API"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'cho_xac_nhan':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chá» cÃ³ thá» xÃ¡c nháº­n booking Äang chá»!'
                })
            
            booking.trang_thai = 'da_xac_nhan'
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ xÃ¡c nháº­n lá»ch háº¹n thÃ nh cÃ´ng!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - api_booking_checkin
@require_role(['nhan_vien', 'quan_ly'])
def api_booking_checkin(request, booking_id):
    """Check-in booking via API"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_xac_nhan':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chá» cÃ³ thá» check-in booking ÄÃ£ xÃ¡c nháº­n!'
                })
            
            booking.trang_thai = 'da_checkin'
            booking.ngay_check_in = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ check-in khÃ¡ch hÃ ng thÃ nh cÃ´ng!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - api_booking_complete_today
@require_role(['nhan_vien', 'quan_ly'])
def api_booking_complete_today(request, booking_id):
    """Complete booking via API for today bookings"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_checkin':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chá» cÃ³ thá» hoÃ n thÃ nh booking ÄÃ£ check-in!'
                })
            
            booking.trang_thai = 'hoan_thanh'
            booking.ngay_hoan_thanh = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ hoÃ n thÃ nh dá»ch vá»¥!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - api_booking_cancel
@require_role(['nhan_vien', 'quan_ly'])
def api_booking_cancel(request, booking_id):
    """Cancel booking via API"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai in ['hoan_thanh', 'da_huy']:
                return JsonResponse({
                    'success': False, 
                    'message': 'KhÃ´ng thá» há»§y booking ÄÃ£ hoÃ n thÃ nh hoáº·c ÄÃ£ há»§y!'
                })
            
            booking.trang_thai = 'da_huy'
            booking.ly_do_huy = request.POST.get('ly_do_huy', '')
            booking.ngay_huy = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ há»§y lá»ch háº¹n!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - api_booking_detail
@require_role(['nhan_vien', 'quan_ly'])
def api_booking_detail(request, booking_id):
    """Get booking detail via API"""
    if request.method == 'GET':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            # Get services for this booking
            services = []
            for dv_dat_lich in booking.dich_vu_dat_lich.all():
                services.append({
                    'ten_dich_vu': dv_dat_lich.dich_vu.ten_dich_vu,
                    'thoi_luong': dv_dat_lich.dich_vu.thoi_gian_thuc_hien,
                    'gia': float(dv_dat_lich.gia_tai_thoi_diem)
                })
            
            data = {
                'id': booking.id,
                'ma_dat_lich': booking.ma_dat_lich,
                'ngay_hen': booking.ngay_hen.strftime('%d/%m/%Y'),
                'gio_bat_dau': booking.gio_hen.strftime('%H:%M'),
                'tong_tien': float(booking.thanh_tien),
                'trang_thai': booking.trang_thai,
                'ghi_chu': booking.ghi_chu or '',
                'khach_hang': {
                    'ho_ten': booking.ten_khach_hang or (booking.khach_hang.ho_ten if booking.khach_hang else ''),
                    'so_dien_thoai': booking.so_dien_thoai_khach,
                    'email': booking.email_khach or (booking.khach_hang.email if booking.khach_hang else '')
                },
                'dich_vu': services
            }
            
            return JsonResponse(data)
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - staff_booking_checkin (staff action)
@require_role(['nhan_vien', 'quan_ly'])
def staff_booking_checkin(request, booking_id):
    """Staff check-in customer"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_xac_nhan':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chá» cÃ³ thá» check-in booking ÄÃ£ xÃ¡c nháº­n!'
                })
            
            booking.trang_thai = 'da_checkin'
            booking.ngay_check_in = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ check-in khÃ¡ch hÃ ng thÃ nh cÃ´ng!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
# - staff_booking_complete (staff action)
@require_role(['nhan_vien', 'quan_ly'])
def staff_booking_complete(request, booking_id):
    """Staff complete booking"""
    if request.method == 'POST':
        try:
            user_id = request.session.get('user_id')
            booking = get_object_or_404(DatLich, id=booking_id, nhan_vien_id=user_id, da_xoa=False)
            
            if booking.trang_thai != 'da_checkin':
                return JsonResponse({
                    'success': False, 
                    'message': 'Chá» cÃ³ thá» hoÃ n thÃ nh booking ÄÃ£ check-in!'
                })
            
            booking.trang_thai = 'hoan_thanh'
            booking.ngay_hoan_thanh = timezone.now()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'ÄÃ£ hoÃ n thÃ nh dá»ch vá»¥!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'CÃ³ lá»i xáº£y ra: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})
